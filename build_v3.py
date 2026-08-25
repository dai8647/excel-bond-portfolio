# -*- coding: utf-8 -*-
"""Capula型 債券RVモデル v3 — ファンド管理版
履歴データ=唯一のデータ源。全シート数式参照。VBA更新対応。BDPなし。
"""
import datetime as dt
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

from master_hist import build_master
from dv01 import par_mod_dur, dv01_per_1mm

OUT = r"C:\Users\dai86\Downloads\Capula_BondRV_Model.xlsx"

NAV_BASE = 100.0  # 仮NAV基数 $mm

# ---------------- スタイル ----------------
NAVY="1F3864"; STEEL="4472C4"; LIGHT="D9E1F2"; GREY="F2F2F2"
GREEN="548235"; RED="C00000"
HDR_FILL=PatternFill("solid",fgColor=NAVY)
SUB_FILL=PatternFill("solid",fgColor=STEEL)
ALT_FILL=PatternFill("solid",fgColor=GREY)
INPUT_FILL=PatternFill("solid",fgColor="FFF2CC")
ASSUMP_FILL=PatternFill("solid",fgColor="FCE4D6")
OK_FILL=PatternFill("solid",fgColor="C6EFCE")
BAD_FILL=PatternFill("solid",fgColor="FFC7CE")
HDR_FONT=Font(color="FFFFFF",bold=True,size=11)
TITLE_FONT=Font(color=NAVY,bold=True,size=16)
SUB_FONT=Font(color=NAVY,bold=True,size=12)
BOLD=Font(bold=True)
NOTE_FONT=Font(italic=True,size=9,color="808080")
THIN=Side(style="thin",color="BFBFBF")
BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
TOPLEFT=Alignment(horizontal="left",vertical="top",wrap_text=True)

def style_header(ws,row,c1,c2,fill=HDR_FILL,font=HDR_FONT):
    for c in range(c1,c2+1):
        cell=ws.cell(row=row,column=c)
        cell.fill=fill;cell.font=font;cell.alignment=CENTER;cell.border=BORDER

def set_widths(ws,widths):
    for i,w in enumerate(widths,start=1):
        ws.column_dimensions[get_column_letter(i)].width=w

def title_block(ws,title,subtitle,ncols):
    ws.cell(row=1,column=1,value=title).font=TITLE_FONT
    ws.cell(row=2,column=1,value=subtitle).font=Font(italic=True,color="808080",size=10)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)

# ---------------- データ ----------------
rows,UST_T,JGB_T=build_master()
last=rows[-1]
UST_Y={t:last[f"UST_{t}"] for t in UST_T}
JGB_Y={t:last[f"JGB_{t}"] for t in JGB_T}
BUND_Y=last["BUND_FUT"]   # Bund 10Y先物インプライド利回り(%)。スポットBundの代用
JGBF_Y=last["JGB_FUT"]    # JGB 10Y先物インプライド利回り(%)。2510 ETF換算
SOFR=last["SOFR"];EFFR=last["EFFR"];TGCR=last["TGCR"];BGCR=last["BGCR"]
USDJPY=last["USDJPY"];EURUSD=last["EURUSD"];ZN=last["ZN"]
ASOF=last["date"]
NROWS=len(rows)          # データ行数
LASTROW=NROWS+1          # シート上の最終データ行(1行目がヘッダ)
# ---- 容量ベース設計: VBA更新で行数が変わっても耐えるよう、エンジンは ----
# ---- 固定容量行まで数式を生成し、空行はIFガードで空白にする。 ----
CAPACITY=NROWS+400       # 履歴データ最大行数(現在+約1.5年分の余裕)
CAP_LAST=CAPACITY+1      # エンジンの最終行(1行目ヘッダ)

def dv01mm(market,tenor):
    if market=="UST":
        y=UST_Y[tenor]
    elif market=="JGB":
        y=JGB_Y[tenor]
    elif market=="BUND":
        y=BUND_Y
    else:
        y=JGBF_Y
    n=float(tenor.rstrip("Y"))
    return dv01_per_1mm(par_mod_dur(n,y))

d2=dv01mm("UST","2Y");d5=dv01mm("UST","5Y");d10=dv01mm("UST","10Y");d30=dv01mm("UST","30Y")
j2=dv01mm("JGB","2Y");j10=dv01mm("JGB","10Y");j30=dv01mm("JGB","30Y")
b10=dv01mm("BUND","10Y");g10=dv01mm("JGBF","10Y")

# 戦略サイズ(DV01 $/bp)
S1_DV01=80000.0; S2_DV01=60000.0; S3_DV01=50000.0; S4_DV01=40000.0
S5_NOT=120.0;    S6_DV01=40000.0
S7_DV01=40000.0; S8_DV01=40000.0
ZN_DV01=79.0

# キャリー(年率$) — 簡易: 現在利回り×レポ想定からビルド時に計算し定数化
GC_USD=TGCR; GC_JPY=-0.10; GC_EUR=1.75   # EUR GCレポはESTR近傍(目分量・ファンディングシートで編集可)
def carry_leg(y,repo,direction):
    return (y-repo) if direction>0 else (repo-y)
CARRY_S1 = S1_DV01/d10*carry_leg(UST_Y["10Y"],GC_USD,+1)*100 + S1_DV01/2/d2*carry_leg(UST_Y["2Y"],GC_USD,-1)*100 + S1_DV01/2/d30*carry_leg(UST_Y["30Y"],GC_USD,-1)*100
CARRY_S2 = S2_DV01/j10*carry_leg(JGB_Y["10Y"],GC_JPY,-1)*100 + S2_DV01/2/j2*carry_leg(JGB_Y["2Y"],GC_JPY,+1)*100 + S2_DV01/2/j30*carry_leg(JGB_Y["30Y"],GC_JPY,+1)*100
CARRY_S3 = S3_DV01/j10*carry_leg(JGB_Y["10Y"],GC_JPY,+1)*100 + S3_DV01/d10*carry_leg(UST_Y["10Y"],GC_USD,-1)*100
CARRY_S4 = S4_DV01/d5*carry_leg(UST_Y["5Y"],GC_USD,+1)*100 + S4_DV01/d10*carry_leg(UST_Y["10Y"],GC_USD,-1)*100
CARRY_S5 = 0.0
CARRY_S6 = S6_DV01/d10*carry_leg(UST_Y["10Y"],GC_USD,+1)*100  # 現物レッグのみ簡易
CARRY_S7 = S7_DV01/b10*carry_leg(BUND_Y,GC_EUR,+1)*100 + S7_DV01/d10*carry_leg(UST_Y["10Y"],GC_USD,-1)*100
CARRY_S8 = S8_DV01/g10*carry_leg(JGBF_Y,GC_JPY,+1)*100 + S8_DV01/b10*carry_leg(BUND_Y,GC_EUR,-1)*100
CARRY_TOTAL=CARRY_S1+CARRY_S2+CARRY_S3+CARRY_S4+CARRY_S5+CARRY_S6+CARRY_S7+CARRY_S8

wb=Workbook()

# =====================================================================
# 履歴データ(唯一のデータ源 / VBAがここを更新)
# =====================================================================
wsH=wb.active;wsH.title="履歴データ"
wsH.sheet_view.showGridLines=False
hist_cols=["日付","UST_2Y","UST_5Y","UST_10Y","UST_20Y","UST_30Y",
           "JGB_2Y","JGB_10Y","JGB_20Y","JGB_30Y",
           "SOFR","EFFR","TGCR","BGCR","USDJPY","EURUSD","ZN先物","BUND_FUT","JGB_FUT"]
wsH.append(hist_cols)
style_header(wsH,1,1,len(hist_cols))
for i,rec in enumerate(rows,start=2):
    d=rec["date"]
    if isinstance(d,str):
        d=dt.date.fromisoformat(d)
    dc=wsH.cell(row=i,column=1,value=d)
    dc.number_format="yyyy-mm-dd"
    vals=[rec.get("UST_2Y"),rec.get("UST_5Y"),rec.get("UST_10Y"),rec.get("UST_20Y"),rec.get("UST_30Y"),
          rec.get("JGB_2Y"),rec.get("JGB_10Y"),rec.get("JGB_20Y"),rec.get("JGB_30Y"),
          rec.get("SOFR"),rec.get("EFFR"),rec.get("TGCR"),rec.get("BGCR"),
          rec.get("USDJPY"),rec.get("EURUSD"),rec.get("ZN"),
          rec.get("BUND_FUT"),rec.get("JGB_FUT")]
    for j,v in enumerate(vals,start=2):
        if v is not None:
            c=wsH.cell(row=i,column=j,value=v)
            c.number_format="0.000" if j not in (16,17) else "0.00"
for i,w in enumerate([12]+[9.5]*(len(hist_cols)-1),start=1):
    wsH.column_dimensions[get_column_letter(i)].width=w
wsH.freeze_panes="B2"
wsH.cell(row=LASTROW+2,column=1,
    value="★このシートが全モデルの唯一のデータ源。ダッシュボードの【データ更新】ボタン(VBA)が米財務省/日本財務省/NY連銀/Yahoo/Pensfordからヒストリカル込みで再取得し、ここを書き換える。書き換え後は全シートが自動再計算。").font=Font(bold=True,color=RED,size=9)
wsH.cell(row=LASTROW+3,column=1,
    value="出典: 米国財務省(daily treasury yield curve)/日本財務省(国債金利情報,日本の休日は前方補完)/NY連銀(SOFR・EFFR・TGCR・BGCR)/Yahoo(USDJPY・EURUSD・ZN先物)/Eurex RX1先物(BUND_FUTインプライド利回り)/2510 ETF(JGB_FUT)。ASWスワップはPensford live-rates(ASWデータシート参照)。").font=NOTE_FONT

# 分析データで使う列文字
HC={"date":"A","UST2":"B","UST5":"C","UST10":"D","UST20":"E","UST30":"F",
    "JGB2":"G","JGB10":"H","JGB20":"I","JGB30":"J",
    "SOFR":"K","EFFR":"L","TGCR":"M","BGCR":"N","USDJPY":"O","EURUSD":"P","ZN":"Q",
    "BUND_FUT":"R","JGB_FUT":"S"}
def H(col,row):
    return f"履歴データ!${HC[col]}${row}"
def Hcol(col):
    return f"履歴データ!${HC[col]}:${HC[col]}"

# 最新行を引く共通式
M_LATEST="MATCH(MAX(履歴データ!$A:$A),履歴データ!$A:$A,0)"
def LATEST(col):
    return f"INDEX({Hcol(col)},{M_LATEST})"

print("履歴データ行:",NROWS,"最終日:",ASOF)
print("キャリー年率$: S1",round(CARRY_S1),"S2",round(CARRY_S2),"S3",round(CARRY_S3),"S4",round(CARRY_S4),"S6",round(CARRY_S6),"S7",round(CARRY_S7),"S8",round(CARRY_S8),"合計",round(CARRY_TOTAL))

# =====================================================================
# 分析データ(数式エンジン: スプレッド→戦略別P&L→仮NAV)
# 列構成: A日付 B S1 C S2 D S3 E S4 F S7 G S8 H USDJPY I ZN J UST10%
#         K..S Δ列(9本)  T..AA 戦略損益(8本)  ABキャリー ACブック損益 AD mm
#         AE仮NAV AFリターン AGピーク AH DD%   (36/37列はS1バンド)
# =====================================================================
wsA=wb.create_sheet("分析データ")
wsA.sheet_view.showGridLines=False
bn_fut=S6_DV01/ZN_DV01
an_hdr=["日付","S1 BF(bp)","S2 JGB BF(bp)","S3 米日ベーシス(bp)","S4 5s10s(bp)","S7 Bund-UST(bp)","S8 JGB-Bund(bp)","USDJPY","ZN先物","UST10%",
        "ΔS1","ΔS2","ΔS3","ΔS4","ΔS7","ΔS8","ΔUSDJPY","ΔZN","ΔUST10(bp)",
        "S1損益$","S2損益$","S3損益$","S4損益$","S5損益$","S6損益$","S7損益$","S8損益$","日次キャリー$","ブック損益$","ブック損益$mm","仮NAV($mm)","日次リターン%","NAVピーク","DD%"]
wsA.append(an_hdr)
style_header(wsA,1,1,len(an_hdr))
for i,w in enumerate([12]+[11]*(len(an_hdr)-1),start=1):
    wsA.column_dimensions[get_column_letter(i)].width=w

# 行r(データは2行目開始)。容量ベース: CAP_LAST行まで数式を生成し、
# 履歴データが行増減してもIFガードで耐える。空行は空白/0になりNAVは横ばい。
carry_daily=CARRY_TOTAL/252
def hd(r):   # 履歴データの日付セル(行r)
    return H('date',r)
for r in range(2,CAP_LAST+1):
    g=hd(r)                      # 当日の日付参照
    gp=hd(r-1) if r>2 else None  # 前日
    nodata=f'{g}=""'
    # 日付・スプレッド(当日データが無ければ空白)
    wsA.cell(row=r,column=1,value=f'=IF({nodata},"",{g})')
    wsA.cell(row=r,column=2,value=f'=IF({nodata},"",({H("UST10",r)}-({H("UST2",r)}+{H("UST30",r)})/2)*100)')
    wsA.cell(row=r,column=3,value=f'=IF({nodata},"",({H("JGB10",r)}-({H("JGB2",r)}+{H("JGB30",r)})/2)*100)')
    wsA.cell(row=r,column=4,value=f'=IF({nodata},"",({H("UST10",r)}-{H("JGB10",r)})*100)')
    wsA.cell(row=r,column=5,value=f'=IF({nodata},"",({H("UST10",r)}-{H("UST5",r)})*100)')
    wsA.cell(row=r,column=6,value=f'=IF({nodata},"",({H("BUND_FUT",r)}-{H("UST10",r)})*100)')
    wsA.cell(row=r,column=7,value=f'=IF({nodata},"",({H("JGB_FUT",r)}-{H("BUND_FUT",r)})*100)')
    wsA.cell(row=r,column=8,value=f'=IF({nodata},"",{H("USDJPY",r)})')
    wsA.cell(row=r,column=9,value=f'=IF({nodata},"",{H("ZN",r)})')
    wsA.cell(row=r,column=10,value=f'=IF({nodata},"",{H("UST10",r)})')
    if r==2:
        # 初回行: Δなし、P&L=0、NAV=基数
        for c in range(11,20):wsA.cell(row=r,column=c,value=None)
        for c in range(20,28):wsA.cell(row=r,column=c,value=0)
        wsA.cell(row=r,column=28,value=f'=IF({nodata},0,{carry_daily:.1f})')
        wsA.cell(row=r,column=29,value=f'=SUM(T{r}:AA{r})')
        wsA.cell(row=r,column=30,value=f'=IF({nodata},"",AC{r}/1000000)')
        wsA.cell(row=r,column=31,value=NAV_BASE)
        wsA.cell(row=r,column=32,value=0)
        wsA.cell(row=r,column=33,value=NAV_BASE)
        wsA.cell(row=r,column=34,value=0)
    else:
        p=r-1
        both=f'OR({nodata},{gp}="")'
        wsA.cell(row=r,column=11,value=f'=IF({both},"",B{r}-B{p})')
        wsA.cell(row=r,column=12,value=f'=IF({both},"",C{r}-C{p})')
        wsA.cell(row=r,column=13,value=f'=IF({both},"",D{r}-D{p})')
        wsA.cell(row=r,column=14,value=f'=IF({both},"",E{r}-E{p})')
        wsA.cell(row=r,column=15,value=f'=IF({both},"",F{r}-F{p})')
        wsA.cell(row=r,column=16,value=f'=IF({both},"",G{r}-G{p})')
        wsA.cell(row=r,column=17,value=f'=IF({both},"",H{r}-H{p})')
        wsA.cell(row=r,column=18,value=f'=IF({both},"",I{r}-I{p})')
        wsA.cell(row=r,column=19,value=f'=IF({both},"",(J{r}-J{p})*100)')
        # 戦略別損益$(空行は"" → COUNT/PERCENTILE/AVERAGEが実データ行のみ集計)
        wsA.cell(row=r,column=20,value=f'=IF({both},"",-{S1_DV01}*K{r})')
        wsA.cell(row=r,column=21,value=f'=IF({both},"",{S2_DV01}*L{r})')
        wsA.cell(row=r,column=22,value=f'=IF({both},"",-{S3_DV01}*M{r})')
        wsA.cell(row=r,column=23,value=f'=IF({both},"",-{S4_DV01}*N{r})')
        wsA.cell(row=r,column=24,value=f'=IF({both},"",{S5_NOT}*1000000*(-Q{r}/H{p}))')
        wsA.cell(row=r,column=25,value=f'=IF({both},"",{S6_DV01}*(-S{r})+(-{bn_fut:.1f})*R{r}*1000)')
        wsA.cell(row=r,column=26,value=f'=IF({both},"",-{S7_DV01}*O{r})')
        wsA.cell(row=r,column=27,value=f'=IF({both},"",-{S8_DV01}*P{r})')
        wsA.cell(row=r,column=28,value=f'=IF({nodata},"",{carry_daily:.1f})')
        wsA.cell(row=r,column=29,value=f'=IF({nodata},"",SUM(T{r}:AA{r}))')
        wsA.cell(row=r,column=30,value=f'=IF({nodata},"",AC{r}/1000000)')
        wsA.cell(row=r,column=31,value=f'=IF({nodata},AE{p},AE{p}+AC{r}/1000000)')
        wsA.cell(row=r,column=32,value=f'=IF(OR({nodata},AE{p}=0),"",(AE{r}-AE{p})/AE{p})')
        wsA.cell(row=r,column=33,value=f'=MAX(AG{p},AE{r})')
        wsA.cell(row=r,column=34,value=f'=IF(OR({nodata},AG{r}=0),"",AE{r}/AG{r}-1)')
# 書式(容量行すべてに適用)
for r in range(2,CAP_LAST+1):
    wsA.cell(row=r,column=1).number_format="yyyy-mm-dd"
    for c in [2,3,4,5,6,7,11,12,13,14,15,16,19]:wsA.cell(row=r,column=c).number_format="0.0"
    for c in [8,9,10,17,18]:wsA.cell(row=r,column=c).number_format="0.00"
    for c in range(20,29):wsA.cell(row=r,column=c).number_format="#,##0"
    wsA.cell(row=r,column=30).number_format="0.00"
    wsA.cell(row=r,column=31).number_format="0.00"
    wsA.cell(row=r,column=32).number_format="0.00%"
    wsA.cell(row=r,column=33).number_format="0.00"
    wsA.cell(row=r,column=34).number_format="0.00%"
wsA.freeze_panes="B2"
# 注記はエンジン容量行(CAP_LAST)の下に置く — データ更新で履歴行が伸びても衝突しない
wsA.cell(row=CAP_LAST+1,column=1,
    value="★各戦略の日次損益をスプレッド変化×DV01で算出し、キャリーを足して仮NAVを構築。S4はスワップの日次履歴が無料では取得できないため5s10sスロープで代用(現在のASW水準は『ASW』シートを参照。BBGで履歴取得後に差し替え)。S6は現物(利回り)+先物(価格)の両レッグでベーシスを評価。S7/S8は先物インプライド利回りによるクロスマーケット(それぞれBund-UST / JGB-Bund)。").font=NOTE_FONT
print("分析データ 構築完了")

# =====================================================================
# エントリーシグナル(Zスコア + ±2σバンド + チャート)
# =====================================================================
wsS=wb.create_sheet("エントリーシグナル")
wsS.sheet_view.showGridLines=False
set_widths(wsS,[3,26,14,14,14,12,12,12,12,16,40])
title_block(wsS,"エントリーシグナル — いつポジションを建てるか","各スプレッドのZスコア(過去1年)で±2σをエントリー圏と判定。チャートはバンド付き。",11)

# 各戦略のスプレッド列(分析データ)と方向
sig_defs=[
 ("S1 UST 2s10s30s BF","分析データ!$B:$B","B",+1,"スプレッド(10年-平均ウィング)が+2σ超=割高→ショートBF(ベリー売り)で参入、-2σ=割安→ロングBF"),
 ("S2 JGB 2s10s30s BF","分析データ!$C:$C","C",+1,"JGBバタフライ。+2σ=割高→ベリー売り、-2σ=割安→ベリー買いで参入"),
 ("S3 UST-JGB 10年 ベーシス","分析データ!$D:$D","D",-1,"ベーシス+2σ=拡大しすぎ→縮小狙い(JGB買い/UST売り)で参入"),
 ("S4 UST 5s10s スロープ","分析データ!$E:$E","E",+1,"5s10sスロープの±2σで参入(ASW水準は『ASW』シート参照)"),
 ("S7 Bund-UST 10年 ベーシス","分析データ!$F:$F","F",-1,"Bund先物インプライド−UST10。+2σ=Bund割高→縮小狙い(Bund売り/UST買い)で参入"),
 ("S8 JGB-Bund 10年 ベーシス","分析データ!$G:$G","G",-1,"JGB先物インプライド−Bundインプライド。+2σ=JGB割高→縮小狙い(JGB売り/Bund買い)で参入"),
]
wsS.append([])
wsS.append(["","戦略","現在値(bp)","1年平均(bp)","1年σ(bp)","Zスコア","+2σ","-2σ","シグナル","推奨アクション",""])
style_header(wsS,3,1,10)
LOOKBACK=252
r=4
sig_rows={}
for name,colref,col,dirn,action in sig_defs:
    cur=f"INDEX({colref},{M_LATEST})"
    mean=f"AVERAGE(OFFSET(分析データ!${col}$1,{M_LATEST}-{LOOKBACK},0,{LOOKBACK},1))"
    std=f"STDEV(OFFSET(分析データ!${col}$1,{M_LATEST}-{LOOKBACK},0,{LOOKBACK},1))"
    z=f"=IFERROR(({cur}-{mean})/{std},0)"
    wsS.cell(row=r,column=2,value=name).font=BOLD
    wsS.cell(row=r,column=3,value=f"={cur}").number_format="0.0"
    wsS.cell(row=r,column=4,value=f"={mean}").number_format="0.0"
    wsS.cell(row=r,column=5,value=f"={std}").number_format="0.0"
    zc=wsS.cell(row=r,column=6,value=z);zc.number_format="0.00";zc.font=BOLD
    wsS.cell(row=r,column=7,value=f"={mean}+2*{std}").number_format="0.0"
    wsS.cell(row=r,column=8,value=f"={mean}-2*{std}").number_format="0.0"
    sig=f'=IF(ABS(F{r})>=2,"★エントリー圏",IF(ABS(F{r})>=1,"様子見","中立"))'
    sc=wsS.cell(row=r,column=9,value=sig)
    wsS.cell(row=r,column=10,value=action).font=Font(size=9)
    wsS.cell(row=r,column=10).alignment=LEFT
    for c in range(2,11):wsS.cell(row=r,column=c).border=BORDER
    sig_rows[name]=r
    r+=1
wsS.cell(row=r+1,column=2,value="※Zスコア=(現在値-過去1年平均)/過去1年σ。|Z|≥2をエントリー圏(平均回帰狙い)とする。実運用ではこれにキャリー・ファンディング・イベントリスクを重ねて最終判断。").font=NOTE_FONT
wsS.cell(row=r+2,column=2,value="※S4はスワップの日次履歴が無料では取得できないため5s10sスロープで運用。現在のASW水準は『ASW』シート、履歴はBBGでスワップレート取得後に差し替え。").font=NOTE_FONT

# チャート: 各スプレッドの履歴 + ±2σバンド(分析データにバンド列を追加)
# バンド列を分析データに追加(列36,37 = S1の+2σ,-2σ 等)。簡易にS1のみチャート例示。
# 分析データにS1バンド列を追加
wsA.cell(row=1,column=36,value="S1+2σ");wsA.cell(row=1,column=37,value="S1-2σ")
for r2 in range(2,LASTROW+1):
    _avg=f"AVERAGE(OFFSET($B$1,MAX(1,ROW()-{LOOKBACK}),0,MIN({LOOKBACK},ROW()-1),1))"
    _sd=f"STDEV(OFFSET($B$1,MAX(1,ROW()-{LOOKBACK}),0,MIN({LOOKBACK},ROW()-1),1))"
    wsA.cell(row=r2,column=36,value=f'=IF(ROW()-1<2,"",IFERROR({_avg}+2*{_sd},""))')
    wsA.cell(row=r2,column=37,value=f'=IF(ROW()-1<2,"",IFERROR({_avg}-2*{_sd},""))')
    wsA.cell(row=r2,column=36).number_format="0.0"
    wsA.cell(row=r2,column=37).number_format="0.0"

ch=LineChart()
ch.title="S1 UST 2s10s30s バタフライスプレッド(±2σバンド)"
ch.style=2;ch.height=9;ch.width=24
ch.y_axis.title="bp";ch.x_axis.title="日"
data=Reference(wsA,min_col=2,min_row=1,max_row=LASTROW)
ch.add_data(data,titles_from_data=True)
band1=Reference(wsA,min_col=36,min_row=1,max_row=LASTROW)
band2=Reference(wsA,min_col=37,min_row=1,max_row=LASTROW)
ch.add_data(band1,titles_from_data=True)
ch.add_data(band2,titles_from_data=True)
cats=Reference(wsA,min_col=1,min_row=2,max_row=LASTROW)
ch.set_categories(cats)
wsS.add_chart(ch,"M3")
print("エントリーシグナル 構築完了")

# =====================================================================
# シグナル検証(±2σエントリー・n日保有バックテスト)
# =====================================================================
wsG=wb.create_sheet("シグナル検証")
wsG.sheet_view.showGridLines=False
set_widths(wsG,[3,26,14,14,14,15,15,14,12])
title_block(wsG,"シグナル検証 — ±2σエントリールールのバックテスト",
    "各戦略スプレッドの252日Zスコアが±閾値を超えた日にエントリー(平均回帰)、n営業日後クローズ。分析データ参照のため更新後自動再計算。",9)

wsG.cell(row=4,column=2,value="■ パラメータ(橙=編集可)").font=SUB_FONT
wsG.cell(row=5,column=2,value="保有期間(営業日)")
wsG.cell(row=5,column=3,value=10).fill=ASSUMP_FILL
wsG.cell(row=6,column=2,value="エントリー閾値(σ)")
wsG.cell(row=6,column=3,value=2).fill=ASSUMP_FILL
wsG.cell(row=7,column=2,value="Zスコア・ルックバック(日)")
wsG.cell(row=7,column=3,value=252).fill=ASSUMP_FILL
for rr in (5,6,7):wsG.cell(row=rr,column=3).border=BORDER

sig_hdr=["戦略","シグナル回数","勝率","平均損益($)","合計損益($)","最大単回損失($)","平均保有(bp)","判定"]
for i,h in enumerate(sig_hdr):wsG.cell(row=9,column=2+i,value=h)
style_header(wsG,9,2,9)

sig_strats=[
 ("S1 UST 2s10s30s BF","B",S1_DV01),
 ("S2 JGB 2s10s30s BF","C",S2_DV01),
 ("S3 UST-JGB 10年ベーシス","D",S3_DV01),
 ("S4 UST 5s10sスロープ","E",S4_DV01),
 ("S7 Bund-UST 10年ベーシス","F",S7_DV01),
 ("S8 JGB-Bund 10年ベーシス","G",S8_DV01),
]

# ヘルパー列レイアウト(右側非表示): 戦略ごとに z/sig/trade/cum/peak/dd、最後に集約4列+日付
HSTART=46
hcols={};trade_cols=[];hb=HSTART
for _n,_L,_V in sig_strats:
    hcols[_L]=dict(z=hb,sig=hb+1,trade=hb+2,cum=hb+3,peak=hb+4,dd=hb+5)
    trade_cols.append(get_column_letter(hb+2));hb+=6
AG_D=hb;AG_C=hb+1;AG_P=hb+2;AG_DD=hb+3;DATE_H=hb+4;LASTH=DATE_H
hL={_L:{k:get_column_letter(v) for k,v in hcols[_L].items()} for _n,_L,_V in sig_strats}
AGDl=get_column_letter(AG_D);AGCl=get_column_letter(AG_C);AGPl=get_column_letter(AG_P);AGDDl=get_column_letter(AG_DD)

for idx,(_n,_L,_V) in enumerate(sig_strats):
    sr=10+idx;tc=hL[_L]["trade"];rng=f"{tc}$2:{tc}${CAP_LAST}"
    wsG.cell(row=sr,column=2,value=_n)
    wsG.cell(row=sr,column=3,value=f"=COUNT({rng})").number_format="0"
    wsG.cell(row=sr,column=4,value=f'=IFERROR(COUNTIF({rng},">0")/COUNT({rng}),"")').number_format="0.0%"
    wsG.cell(row=sr,column=5,value=f'=IFERROR(AVERAGE({rng}),"")').number_format="#,##0"
    wsG.cell(row=sr,column=6,value=f"=SUM({rng})").number_format="#,##0"
    wsG.cell(row=sr,column=7,value=f'=IFERROR(MIN({rng}),"")').number_format="#,##0"
    wsG.cell(row=sr,column=8,value=f'=IFERROR(AVERAGE({rng})/{_V:.0f},"")').number_format="0.0"
    wsG.cell(row=sr,column=9,value=f'=IF(COUNT({rng})=0,"データ不足",IF(AND(F{sr}>0,D{sr}>=0.5),"採用候補",IF(F{sr}>0,"要観察","見送り")))')
    for c in range(2,10):wsG.cell(row=sr,column=c).border=BORDER

num_win="+".join(f'COUNTIF({t}$2:{t}${CAP_LAST},">0")' for t in trade_cols)
min_all=",".join(f"{t}$2:{t}${CAP_LAST}" for t in trade_cols)
wsG.cell(row=16,column=2,value="全戦略合計").font=BOLD
wsG.cell(row=16,column=3,value="=SUM(C10:C15)").number_format="0"
wsG.cell(row=16,column=4,value=f'=IFERROR(({num_win})/C16,"")').number_format="0.0%"
wsG.cell(row=16,column=5,value='=IFERROR(F16/C16,"")').number_format="#,##0"
wsG.cell(row=16,column=6,value="=SUM(F10:F15)").number_format="#,##0"
wsG.cell(row=16,column=7,value=f"=MIN({min_all})").number_format="#,##0"
wsG.cell(row=16,column=9,value='=IF(AND(F16>0,D16>=0.5),"採用候補",IF(F16>0,"要観察","見送り"))')
for c in range(2,10):wsG.cell(row=16,column=c).border=BORDER;wsG.cell(row=16,column=c).font=BOLD

wsG.cell(row=18,column=2,value="■ ポートフォリオ(全戦略合算)のリスク指標").font=SUB_FONT
wsG.cell(row=19,column=2,value="最大ドローダウン($)")
wsG.cell(row=19,column=3,value=f"=MIN({AGDDl}$2:{AGDDl}${CAP_LAST})").number_format="#,##0"
wsG.cell(row=20,column=2,value="最大DD(仮NAV $100mm比 %)")
wsG.cell(row=20,column=3,value="=C19/100000000").number_format="0.00%"
wsG.cell(row=21,column=2,value="累積トレード損益($)")
wsG.cell(row=21,column=3,value="=F16").number_format="#,##0"
wsG.cell(row=22,column=2,value="総トレード回数")
wsG.cell(row=22,column=3,value="=C16").number_format="0"

for i,nt in enumerate([
 "【方法】各戦略スプレッドの252日Zスコアが±閾値σを超えた日にエントリー(+σ=ショート、−σ=ロングの平均回帰)、n営業日後クローズ。",
 "トレード損益 = DV01 × SIGN(Zエントリー) × (エントリー時スプレッド − クローズ時スプレッド)。取引コスト・ポジション上限は未反映。",
 "累積カーブは各トレード損益をエントリー日に合算(トレード重複を許容)した参考値。S4はスワップ履歴が無いため5s10sスロープで代用。",
 "右側の非表示列に日次Zスコア・シグナル・トレード損益・累積カーブを保持。パラメータ(C5:C7)を変えると全て再計算。"]):
    wsG.cell(row=24+i,column=2,value=nt).font=NOTE_FONT

for _n,_L,_V in sig_strats:
    wsG.cell(row=1,column=hcols[_L]["cum"],value=0)
    wsG.cell(row=1,column=hcols[_L]["peak"],value=0)
    wsG.cell(row=1,column=hcols[_L]["dd"],value=0)
wsG.cell(row=1,column=AG_C,value=0);wsG.cell(row=1,column=AG_P,value=0);wsG.cell(row=1,column=AG_DD,value=0)

def _z(_L,r):
    return (f'=IF({r}<($C$7+1),"",IFERROR(('
        f'分析データ!{_L}{r}-AVERAGE(OFFSET(分析データ!{_L}$1,MAX(1,{r}-$C$7),0,MIN($C$7,{r}-1),1))'
        f')/STDEV(OFFSET(分析データ!{_L}$1,MAX(1,{r}-$C$7),0,MIN($C$7,{r}-1),1)),0))')

for r in range(2,CAP_LAST+1):
    for _n,_L,_V in sig_strats:
        zc=hL[_L]["z"];sc=hL[_L]["sig"];tc=hL[_L]["trade"];cc=hL[_L]["cum"];pc=hL[_L]["peak"];dc=hL[_L]["dd"]
        wsG.cell(row=r,column=hcols[_L]["z"],value=_z(_L,r))
        wsG.cell(row=r,column=hcols[_L]["sig"],value=f'=IF(ISNUMBER({zc}{r}),IF(ABS({zc}{r})>=$C$6,1,0),0)')
        wsG.cell(row=r,column=hcols[_L]["trade"],value=(
            f'=IF({sc}{r}=1,IF(INDEX(分析データ!$A:$A,{r}+$C$5)<>"",'
            f'{_V:.0f}*SIGN({zc}{r})*(分析データ!{_L}{r}-INDEX(分析データ!{_L}:{_L},{r}+$C$5)),""),"")'))
        wsG.cell(row=r,column=hcols[_L]["cum"],value=f'={cc}{r-1}+IF(AND({sc}{r}=1,ISNUMBER({tc}{r})),{tc}{r},0)')
        wsG.cell(row=r,column=hcols[_L]["peak"],value=f'=MAX({pc}{r-1},{cc}{r})')
        wsG.cell(row=r,column=hcols[_L]["dd"],value=f'={cc}{r}-{pc}{r}')
    dt_f="+".join(f"N({t}{r})" for t in trade_cols)
    wsG.cell(row=r,column=AG_D,value=f"={dt_f}")
    wsG.cell(row=r,column=AG_C,value=f"={AGCl}{r-1}+{AGDl}{r}")
    wsG.cell(row=r,column=AG_P,value=f"=MAX({AGPl}{r-1},{AGCl}{r})")
    wsG.cell(row=r,column=AG_DD,value=f"={AGCl}{r}-{AGPl}{r}")
    wsG.cell(row=r,column=DATE_H,value=f"=分析データ!A{r}").number_format="yyyy-mm-dd"

for c in range(HSTART,LASTH+1):wsG.column_dimensions[get_column_letter(c)].hidden=True

chG=LineChart();chG.title="シグナル戦略 累積損益(全戦略合算・参考)"
chG.style=2;chG.height=10;chG.width=24;chG.y_axis.title="$";chG.x_axis.title="日"
chG.add_data(Reference(wsG,min_col=AG_C,min_row=1,max_row=CAP_LAST),titles_from_data=False)
chG.set_categories(Reference(wsG,min_col=DATE_H,min_row=2,max_row=CAP_LAST))
wsG.add_chart(chG,"K9")
print("シグナル検証 構築完了")

# =====================================================================
# トレード台帳(オープン評価 + クローズド記入 + バックテスト収益表)
# =====================================================================
wsT=wb.create_sheet("トレード台帳")
wsT.sheet_view.showGridLines=False
set_widths(wsT,[3,26,12,12,12,12,13,13,13,13,13,30])
title_block(wsT,"トレード台帳 — 戦略別収益表","オープンポジションは自動評価(分析データ参照)。クローズドトレードは手動記入。",12)

# --- バックテスト収益表(2023年以降の戦略別累積損益) ---
wsT.append([])
wsT.cell(row=3,column=2,value="■ バックテスト収益表(2023-01〜現在、分析データから自動集計)").font=SUB_FONT
wsT.append([])
hdr=["","戦略","累積マーク損益$","累積キャリー$","合計損益$","勝率(日次)","日数","シャーレシオ(日次)","最大DD$","備考"]
for i,h in enumerate(hdr):wsT.cell(row=5,column=1+i,value=h)
style_header(wsT,5,1,10)
strat_cols={"S1 UST 2s10s30s BF":"T","S2 JGB 2s10s30s BF":"U","S3 UST-JGB 10年 クロスマーケット":"V","S4 UST ASWボックス 5s10s":"W","S5 ドル円オーバーレイ":"X","S6 UST 10年 ベーシス":"Y","S7 Bund-UST 10年 クロスマーケット":"Z","S8 JGB-Bund 10年 クロスマーケット":"AA"}
strat_carry_const={"S1 UST 2s10s30s BF":CARRY_S1,"S2 JGB 2s10s30s BF":CARRY_S2,"S3 UST-JGB 10年 クロスマーケット":CARRY_S3,"S4 UST ASWボックス 5s10s":CARRY_S4,"S5 ドル円オーバーレイ":CARRY_S5,"S6 UST 10年 ベーシス":CARRY_S6,"S7 Bund-UST 10年 クロスマーケット":CARRY_S7,"S8 JGB-Bund 10年 クロスマーケット":CARRY_S8}
r=6
for name,col in strat_cols.items():
    wsT.cell(row=r,column=2,value=name)
    mc=wsT.cell(row=r,column=3,value=f"=SUM(分析データ!${col}$2:${col}${CAP_LAST})");mc.number_format="#,##0"
    cc=wsT.cell(row=r,column=4,value=round(strat_carry_const[name]*(NROWS/252)));cc.number_format="#,##0"
    tc=wsT.cell(row=r,column=5,value=f"=C{r}+D{r}");tc.number_format="#,##0";tc.font=BOLD
    wr=wsT.cell(row=r,column=6,value=f"=COUNTIF(分析データ!${col}$2:${col}${CAP_LAST},\">0\")/COUNT(分析データ!${col}$2:${col}${CAP_LAST})");wr.number_format="0.0%"
    wsT.cell(row=r,column=7,value=NROWS-1).number_format="#,##0"
    sr=wsT.cell(row=r,column=8,value=f"=IFERROR(AVERAGE(分析データ!${col}$2:${col}${CAP_LAST})/STDEV(分析データ!${col}$2:${col}${CAP_LAST})*SQRT(252),0)");sr.number_format="0.00"
    wsT.cell(row=r,column=9,value="").number_format="#,##0"
    wsT.cell(row=r,column=10,value="バックテスト(2023-01〜)").font=Font(size=9,color="595959")
    for c in range(2,11):wsT.cell(row=r,column=c).border=BORDER
    r+=1
# 合計行
wsT.cell(row=r,column=2,value="ブック合計").font=BOLD
wsT.cell(row=r,column=3,value=f"=SUM(C6:C{r-1})").number_format="#,##0"
wsT.cell(row=r,column=4,value=f"=SUM(D6:D{r-1})").number_format="#,##0"
wsT.cell(row=r,column=5,value=f"=SUM(E6:E{r-1})").number_format="#,##0"
for c in range(2,11):wsT.cell(row=r,column=c).fill=PatternFill("solid",fgColor=LIGHT);wsT.cell(row=r,column=c).border=BORDER;wsT.cell(row=r,column=c).font=BOLD
r+=2

# --- オープンポジション(モデルポート) ---
wsT.cell(row=r,column=2,value="■ オープンポジション(モデルポート、現在値評価)").font=SUB_FONT
r+=1
hdr2=["","戦略","建値スプレッド(bp)","現在スプレッド(bp)","含み損益$","DV01($/bp)","方向","エントリー日","状態",""]
for i,h in enumerate(hdr2):wsT.cell(row=r,column=1+i,value=h)
style_header(wsT,r,1,10)
r+=1
open_pos=[
 ("S1 UST 2s10s30s BF","分析データ!$B:$B",+1,S1_DV01,"モデル建玉(バックテスト起点=2023-01)"),
 ("S2 JGB 2s10s30s BF","分析データ!$C:$C",-1,S2_DV01,"モデル建玉"),
 ("S3 UST-JGB 10年 クロスマーケット","分析データ!$D:$D",-1,S3_DV01,"モデル建玉"),
 ("S4 UST ASWボックス 5s10s","分析データ!$E:$E",+1,S4_DV01,"モデル建玉(5s10sスロープ運用・ASW水準は『ASW』シート)"),
 ("S6 UST 10年 ベーシス","分析データ!$I:$I",+1,S6_DV01,"モデル建玉(先物価格ベース)"),
 ("S7 Bund-UST 10年 クロスマーケット","分析データ!$F:$F",-1,S7_DV01,"モデル建玉(先物インプライド)"),
 ("S8 JGB-Bund 10年 クロスマーケット","分析データ!$G:$G",-1,S8_DV01,"モデル建玉(先物インプライド)"),
]
for name,colref,dirn,dv01,note in open_pos:
    cur=f"INDEX({colref},{M_LATEST})"
    wsT.cell(row=r,column=2,value=name)
    wsT.cell(row=r,column=3,value="建値を記入").fill=INPUT_FILL
    wsT.cell(row=r,column=4,value=f"={cur}").number_format="0.0"
    wsT.cell(row=r,column=5,value=f"=IF(ISNUMBER(C{r}),{dirn}*{dv01}*(D{r}-C{r}),\"\")").number_format="#,##0"
    wsT.cell(row=r,column=6,value=dv01).number_format="#,##0"
    wsT.cell(row=r,column=7,value="買い" if dirn>0 else "売り")
    wsT.cell(row=r,column=8,value="").fill=INPUT_FILL
    wsT.cell(row=r,column=9,value="オープン")
    wsT.cell(row=r,column=10,value=note).font=Font(size=9,color="595959")
    for c in range(2,11):wsT.cell(row=r,column=c).border=BORDER
    r+=1
r+=1
# --- クローズドトレード(手動記入) ---
wsT.cell(row=r,column=2,value="■ クローズドトレード(手動記入)").font=SUB_FONT
r+=1
hdr3=["","戦略","エントリー日","クローズ日","建値(bp)","クローズ値(bp)","DV01($/bp)","実現損益$","キャリー$","合計$","メモ"]
for i,h in enumerate(hdr3):wsT.cell(row=r,column=1+i,value=h)
style_header(wsT,r,1,11)
r+=1
for k in range(8):
    wsT.cell(row=r,column=8,value=f"=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r}),ISNUMBER(G{r})),G{r}*(F{r}-E{r}),\"\")").number_format="#,##0"
    wsT.cell(row=r,column=10,value=f"=IF(ISNUMBER(H{r}),H{r}+IF(ISNUMBER(I{r}),I{r},0),\"\")").number_format="#,##0"
    for c in range(2,12):wsT.cell(row=r,column=c).border=BORDER;wsT.cell(row=r,column=c).fill=INPUT_FILL
    r+=1
wsT.cell(row=r+1,column=2,value="※クローズドトレードは黄色セルに記入。実現損益=DV01×(クローズ値-建値)で自動計算。S5(FX)は別途。").font=NOTE_FONT
print("トレード台帳 構築完了")

# =====================================================================
# ダッシュボード(仮NAV・リスク指標・更新ボタン)
# =====================================================================
wsD=wb.create_sheet("ダッシュボード",0)
wsD.sheet_view.showGridLines=False
set_widths(wsD,[3,30,16,16,16,16,16,16])
title_block(wsD,"ブック ダッシュボード","Capula型 債券RV — 仮NAV・リスク指標・戦略サマリ / 基準日 "+ASOF,8)

# 更新ボタン(VBA)
wsD.cell(row=4,column=2,value="▶ データ更新(ヒストリカル取得)").font=Font(bold=True,color="FFFFFF",size=12)
wsD.cell(row=4,column=2).fill=PatternFill("solid",fgColor=GREEN)
wsD.cell(row=4,column=2).alignment=CENTER
wsD.merge_cells("B4:C4")
wsD.cell(row=4,column=2).hyperlink=None
# ボタン注記
wsD.cell(row=5,column=2,value="↑クリックでVBAが米財務省/財務省/NY連銀/Yahoo/Eurex/2510/Pensford(ASW)からヒストリカル込みで再取得し、全シート自動再計算。").font=NOTE_FONT
wsD.merge_cells("B5:H5")

# KPI(仮NAV・リスク)。実データ行数は日付列AのCOUNTで数える(空行はNAVが横ばい転写されるためY列COUNTは不可)。
wsD.cell(row=7,column=2,value="■ パフォーマンス(仮NAV / バックテスト)").font=SUB_FONT
NDAYS=f"COUNT(分析データ!$A$2:$A${CAP_LAST})"
kpis=[
 ("仮NAV($mm)",f"=INDEX(分析データ!$AE:$AE,{M_LATEST})","0.00"),
 ("期間リターン%",f"=INDEX(分析データ!$AE:$AE,{M_LATEST})/{NAV_BASE}-1","0.0%"),
 ("年率リターン%",f"=(INDEX(分析データ!$AE:$AE,{M_LATEST})/{NAV_BASE})^(252/{NDAYS})-1","0.0%"),
 ("シャーレシオ",f"=IFERROR(AVERAGE(分析データ!$AF$2:$AF${CAP_LAST})/STDEV(分析データ!$AF$2:$AF${CAP_LAST})*SQRT(252),0)","0.00"),
 ("最大DD%",f"=MIN(分析データ!$AH$2:$AH${CAP_LAST})","0.0%"),
 ("VaR99%(1日,$mm)",f"=-PERCENTILE(分析データ!$AD$2:$AD${CAP_LAST},0.01)","0.00"),
]
r=8
for i,(label,formula,fmt) in enumerate(kpis):
    col=2+(i%3)*2
    row=r+(i//3)*2
    wsD.cell(row=row,column=col,value=label).font=Font(bold=True,color="FFFFFF",size=9)
    wsD.cell(row=row,column=col).fill=PatternFill("solid",fgColor=STEEL)
    wsD.cell(row=row,column=col).alignment=CENTER
    wsD.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
    v=wsD.cell(row=row+1,column=col,value=formula);v.number_format=fmt
    v.font=Font(bold=True,size=15,color=NAVY);v.alignment=CENTER
    wsD.cell(row=row+1,column=col).fill=PatternFill("solid",fgColor=LIGHT)
    wsD.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
    for rr in (row,row+1):
        for cc in (col,col+1):wsD.cell(row=rr,column=cc).border=BORDER
wsD.row_dimensions[9].height=26;wsD.row_dimensions[11].height=26

# 戦略サマリ(全6戦略を統一フォーマットで)
wsD.cell(row=14,column=2,value="■ 戦略別サマリ(現在値・シグナル)").font=SUB_FONT
r=15
hdr=["","戦略","現在値","Zスコア","シグナル","サイズ","キャリー$/年",""]
for i,h in enumerate(hdr):wsD.cell(row=r,column=1+i,value=h)
style_header(wsD,r,1,8)
r+=1
# (名称, 参照列, 値書式, サイズ表示, サイズ値, キャリー)
dash_strats=[
 ("S1 UST 2s10s30s BF","分析データ!$B:$B","0.0",f"DV01 ${S1_DV01:,.0f}/bp",S1_DV01,CARRY_S1),
 ("S2 JGB 2s10s30s BF","分析データ!$C:$C","0.0",f"DV01 ${S2_DV01:,.0f}/bp",S2_DV01,CARRY_S2),
 ("S3 UST-JGB 10年 クロスマーケット","分析データ!$D:$D","0.0",f"DV01 ${S3_DV01:,.0f}/bp",S3_DV01,CARRY_S3),
 ("S4 UST ASWボックス 5s10s","分析データ!$E:$E","0.0",f"DV01 ${S4_DV01:,.0f}/bp",S4_DV01,CARRY_S4),
 ("S5 ドル円オーバーレイ","分析データ!$H:$H","0.00",f"ノショナル ${S5_NOT:.0f}mm",S5_NOT,CARRY_S5),
 ("S6 UST 10年 ベーシス","分析データ!$I:$I","0.00",f"DV01 ${S6_DV01:,.0f}/bp",S6_DV01,CARRY_S6),
 ("S7 Bund-UST 10年 クロスマーケット","分析データ!$F:$F","0.0",f"DV01 ${S7_DV01:,.0f}/bp",S7_DV01,CARRY_S7),
 ("S8 JGB-Bund 10年 クロスマーケット","分析データ!$G:$G","0.0",f"DV01 ${S8_DV01:,.0f}/bp",S8_DV01,CARRY_S8),
]
for name,colref,fmt,sizelabel,size,carry in dash_strats:
    cur=f"INDEX({colref},{M_LATEST})"
    col=colref.split("$")[1].rstrip(":")
    mean=f"AVERAGE(OFFSET(分析データ!${col}$1,{M_LATEST}-{LOOKBACK},0,{LOOKBACK},1))"
    std=f"STDEV(OFFSET(分析データ!${col}$1,{M_LATEST}-{LOOKBACK},0,{LOOKBACK},1))"
    wsD.cell(row=r,column=2,value=name)
    wsD.cell(row=r,column=3,value=f"={cur}").number_format=fmt
    wsD.cell(row=r,column=4,value=f"=IFERROR(({cur}-{mean})/{std},0)").number_format="0.00"
    wsD.cell(row=r,column=5,value=f'=IF(ABS(D{r})>=2,"★エントリー圏",IF(ABS(D{r})>=1,"様子見","中立"))')
    wsD.cell(row=r,column=6,value=sizelabel).font=Font(size=9)
    cc=wsD.cell(row=r,column=7,value=round(carry));cc.number_format="#,##0"
    cc.font=Font(color=GREEN if carry>0 else RED)
    for c in range(2,8):wsD.cell(row=r,column=c).border=BORDER
    r+=1
wsD.cell(row=r+1,column=2,value="※現在値: S1-S4/S7/S8はスプレッド(bp)、S5はUSDJPY水準、S6はZN先物価格。仮NAVは2023-01からのバックテスト(分析データ)。実運用NAVではない。VaR99は日次損益の1%タイル。").font=NOTE_FONT

# ---- チャート: 仮NAV推移 / ドローダウン / 戦略別損益 ----
from openpyxl.chart import BarChart
# 仮NAV推移
ch_nav=LineChart()
ch_nav.title="仮NAV推移 ($mm) — バックテスト"
ch_nav.style=2;ch_nav.height=8;ch_nav.width=22
ch_nav.y_axis.title="$mm";ch_nav.x_axis.title="日"
nav_data=Reference(wsA,min_col=31,min_row=1,max_row=LASTROW)
ch_nav.add_data(nav_data,titles_from_data=True)
nav_cats=Reference(wsA,min_col=1,min_row=2,max_row=LASTROW)
ch_nav.set_categories(nav_cats)
wsD.add_chart(ch_nav,"B24")
# ドローダウン
ch_dd=LineChart()
ch_dd.title="ドローダウン (%) — NAVピーク比"
ch_dd.style=2;ch_dd.height=8;ch_dd.width=22
ch_dd.y_axis.title="%";ch_dd.x_axis.title="日"
dd_data=Reference(wsA,min_col=34,min_row=1,max_row=LASTROW)
ch_dd.add_data(dd_data,titles_from_data=True)
ch_dd.set_categories(nav_cats)
wsD.add_chart(ch_dd,"B41")
# 戦略別 合計損益(トレード台帳 E6:E14)
ch_attr=BarChart()
ch_attr.type="col";ch_attr.style=10
ch_attr.title="戦略別 合計損益 ($, バックテスト)"
ch_attr.y_axis.title="$";ch_attr.x_axis.title="戦略"
attr_data=Reference(wsT,min_col=5,min_row=5,max_row=13)
ch_attr.add_data(attr_data,titles_from_data=True)
attr_cats=Reference(wsT,min_col=2,min_row=6,max_row=13)
ch_attr.set_categories(attr_cats)
ch_attr.legend=None
wsD.add_chart(ch_attr,"M24")
print("ダッシュボード 構築完了")

# =====================================================================
# イールドカーブ(履歴データ参照 — VBA更新で自動反映)
# =====================================================================
ws=wb.create_sheet("イールドカーブ")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,12,14,14,16])
title_block(ws,"イールドカーブ(最新値 = 履歴データ参照)","VBA更新で自動反映。BBGを使う場合はこのシートの値セルに直接BDPを貼れば上書きできる。",5)
ws.append([]);ws.append(["","年限","米国債%","日本国債%","米-日ベーシス(bp)"])
style_header(ws,3,1,5)
curve_map=[("2Y","UST2","JGB2"),("5Y","UST5",None),("10Y","UST10","JGB10"),("20Y","UST20","JGB20"),("30Y","UST30","JGB30")]
r=4
for t,uk,jk in curve_map:
    ws.cell(row=r,column=2,value=t).font=BOLD
    cu=ws.cell(row=r,column=3,value=f"={LATEST(uk)}");cu.number_format="0.000"
    if jk:
        cj=ws.cell(row=r,column=4,value=f"={LATEST(jk)}");cj.number_format="0.000"
        cb=ws.cell(row=r,column=5,value=f"=(C{r}-D{r})*100");cb.number_format="0.0"
    for c in range(2,6):ws.cell(row=r,column=c).border=BORDER
    r+=1
ws.cell(row=r+1,column=2,value="※履歴データシートの最新行をINDEX/MATCHで参照。VBAが履歴を更新すると自動で変わる。").font=NOTE_FONT

# =====================================================================
# ASWデータ(Pensford live-rates — VBA【データ更新】ボタンが上書きする)
# =====================================================================
wsAswD=wb.create_sheet("ASWデータ")
wsAswD.sheet_view.showGridLines=False
set_widths(wsAswD,[3,22,16,44])
title_block(wsAswD,"ASWデータ — Pensford ライブレート","出典: https://pensford.com/api/live-rates (公開API・動作確認済み)。VBA更新ボタンで上書き。",4)
wsAswD.append([])
wsAswD.cell(row=3,column=2,value="フィールド");wsAswD.cell(row=3,column=3,value="値");wsAswD.cell(row=3,column=4,value="説明")
style_header(wsAswD,3,2,4)
asw_fields=[
 ("quoteDate","", "データ日付(米国)" ),
 ("dailySofr","", "O/N SOFR"),
 ("termSofr1M","", "1M Term SOFR"),
 ("termSofr3M","", "3M Term SOFR"),
 ("fedFunds","", "FFレート"),
 ("treasury2Y","", "米2年利回り"),
 ("treasury3Y","", "米3年利回り"),
 ("treasury5Y","", "米5年利回り"),
 ("treasury7Y","", "米7年利回り"),
 ("treasury10Y","", "米10年利回り"),
 ("oisSwap2Y","", "SOFRスワップ2年"),
 ("oisSwap3Y","", "SOFRスワップ3年"),
 ("oisSwap5Y","", "SOFRスワップ5年"),
 ("oisSwap7Y","", "SOFRスワップ7年"),
 ("oisSwap10Y","", "SOFRスワップ10年"),
 ("bankSwap10Y","", "銀行クレジット込みスワップ10年"),
]
r=4
for f,val,desc in asw_fields:
    wsAswD.cell(row=r,column=2,value=f).font=BOLD
    wsAswD.cell(row=r,column=3,value=val).number_format="0.0000"
    wsAswD.cell(row=r,column=4,value=desc).font=Font(size=9,color="595959")
    for c in range(2,5):wsAswD.cell(row=r,column=c).border=BORDER
    r+=1
wsAswD.cell(row=r+1,column=2,value="※このシートはVBA(データ更新ボタン)がPensfordの公開APIを叩いて上書きする。BBG利用時は=BDH()で差し替えてもOK。").font=NOTE_FONT
print("ASWデータ 構築完了")

# =====================================================================
# ASW(資産スワップスプレッド モニター)
# =====================================================================
wsAsw=wb.create_sheet("ASW")
wsAsw.sheet_view.showGridLines=False
set_widths(wsAsw,[3,26,15,15,15,13,16,42])
title_block(wsAsw,"ASW モニター — 資産スワップスプレッド","債券利回り − スワップレート。米国はPensfordから自動取得。Bund/JGBは橙セルにスワップを手入力(BBGがあれば=BDH()で自動化可)。",8)

# --- UST ASW ---
wsAsw.append([])
wsAsw.cell(row=4,column=2,value="■ 米国債 ASWカーブ(Pensford, ASWデータシート参照)").font=SUB_FONT
hdr=["","年限","米国債利回り%","SOFRスワップ%","ASW(bp)","","備考"]
for i,h in enumerate(hdr):wsAsw.cell(row=5,column=1+i,value=h)
style_header(wsAsw,5,1,7)
# ASWデータシートの行マップ(上記asw_fieldsの並び): 2Y→9/14, 3Y→10/15, 5Y→11/16, 7Y→12/17, 10Y→13/18
asw_row_map={"2Y":(9,14),"3Y":(10,15),"5Y":(11,16),"7Y":(12,17),"10Y":(13,18)}
r=6
for t,(tr_row,sw_row) in asw_row_map.items():
    wsAsw.cell(row=r,column=2,value=t).font=BOLD
    tc=wsAsw.cell(row=r,column=3,value=f'=IF(ISNUMBER(ASWデータ!B{tr_row}),ASWデータ!B{tr_row},"")');tc.number_format="0.000%"
    sc=wsAsw.cell(row=r,column=4,value=f'=IF(ISNUMBER(ASWデータ!B{sw_row}),ASWデータ!B{sw_row},"")');sc.number_format="0.000%"
    ac=wsAsw.cell(row=r,column=5,value=f'=IF(OR(ISNUMBER(ASWデータ!B{tr_row}),ISNUMBER(ASWデータ!B{sw_row})),(ASWデータ!B{tr_row}-ASWデータ!B{sw_row})*10000,"")');ac.number_format="0.0";ac.font=BOLD
    for c in range(2,6):wsAsw.cell(row=r,column=c).border=BORDER
    r+=1
wsAsw.cell(row=r,column=2,value="※ASW=米国債利回り−SOFRスワップ。プラス=現物がスワップより割高(リスクフリー利回り超過)。").font=NOTE_FONT

# --- Bund ASW ---
r+=1
wsAsw.cell(row=r,column=2,value="■ Bund ASW(10年)").font=SUB_FONT
r+=1
wsAsw.cell(row=r,column=2,value="Bund 10年利回り%(先物インプライド)").font=Font(size=9)
bc=wsAsw.cell(row=r,column=3,value=f'={LATEST("BUND_FUT")}');bc.number_format="0.000"
wsAsw.cell(row=r,column=4,value="EUR 10年スワップ%(手入力)").font=Font(size=9)
ec=wsAsw.cell(row=r,column=5,value=2.50);ec.fill=ASSUMP_FILL;ec.number_format="0.000"   # 橙セル
r+=1
wsAsw.cell(row=r,column=2,value="Bund ASW(bp)").font=BOLD
ac=wsAsw.cell(row=r,column=3,value=f'=IF(OR(C{r-1}="",E{r-1}=""),"",(C{r-1}-E{r-1})*100)');ac.number_format="0.0";ac.font=BOLD
wsAsw.cell(row=r,column=4,value="※EURスワップは無料APIが無いため手入力。BBGなら=EURSW10 Curncy 等を貼り付け。").font=Font(size=9,color="595959")

# --- JGB ASW ---
r+=1
wsAsw.cell(row=r,column=2,value="■ JGB ASW(10年)").font=SUB_FONT
r+=1
wsAsw.cell(row=r,column=2,value="JGB 10年利回り%(財務省)").font=Font(size=9)
jc=wsAsw.cell(row=r,column=3,value=f'={LATEST("JGB10")}');jc.number_format="0.000"
wsAsw.cell(row=r,column=4,value="JPY 10年スワップ%(手入力)").font=Font(size=9)
yc=wsAsw.cell(row=r,column=5,value=2.20);yc.fill=ASSUMP_FILL;yc.number_format="0.000"   # 橙セル
r+=1
wsAsw.cell(row=r,column=2,value="JGB ASW(bp)").font=BOLD
ac2=wsAsw.cell(row=r,column=3,value=f'=IF(OR(C{r-1}="",E{r-1}=""),"",(C{r-1}-E{r-1})*100)');ac2.number_format="0.0";ac2.font=BOLD
wsAsw.cell(row=r,column=4,value="※JPYスワップは無料APIが無いため手入力。BBGなら=JPYSW10 Curncy 等を貼り付け。").font=Font(size=9,color="595959")

# --- 先物ベーシス(現先) 参考 ---
r+=1
wsAsw.cell(row=r,column=2,value="■ 先物ベーシス(現物利回り−先物インプライド・参考)").font=SUB_FONT
r+=1
wsAsw.cell(row=r,column=2,value="Bund 現先ベーシス(bp)").font=Font(size=9)
wsAsw.cell(row=r,column=3,value="(スポットBund未取得・先物インプライドを表示)").font=Font(size=9,color="595959")
bc2=wsAsw.cell(row=r,column=4,value=f'={LATEST("BUND_FUT")}');bc2.number_format="0.000"
r+=1
wsAsw.cell(row=r,column=2,value="JGB 現先ベーシス(bp)").font=Font(size=9)
jb=wsAsw.cell(row=r,column=3,value=f'=IF(OR({LATEST("JGB_FUT")}="",{LATEST("JGB10")}=""),"",({LATEST("JGB_FUT")}-{LATEST("JGB10")})*100)');jb.number_format="0.0"
wsAsw.cell(row=r,column=4,value="※プラス=先物インプライドが現物より高い(先物が割安)。").font=Font(size=9,color="595959")

# --- チャート: UST ASWカーブ ---
r+=1
wsAsw.cell(row=r,column=2,value="■ UST ASWカーブ(bp)").font=SUB_FONT
from openpyxl.chart import BarChart
ch_asw=BarChart()
ch_asw.type="col";ch_asw.style=10
ch_asw.title="UST ASWカーブ (bp)"
ch_asw.y_axis.title="bp";ch_asw.x_axis.title="年限"
asw_data=Reference(wsAsw,min_col=5,min_row=5,max_row=10)
ch_asw.add_data(asw_data,titles_from_data=True)
asw_cats=Reference(wsAsw,min_col=2,min_row=6,max_row=10)
ch_asw.set_categories(asw_cats)
ch_asw.legend=None
ch_asw.height=8;ch_asw.width=18
wsAsw.add_chart(ch_asw,"G6")

# --- データソース表 ---
r+=2
wsAsw.cell(row=r,column=2,value="■ データソース(BBGなしで取得できるもの)").font=SUB_FONT
r+=1
for label,src,note in [
 ("米国債・SOFRスワップ", "https://pensford.com/api/live-rates", "Pensford公開API。VBA更新ボタンで自動取得(動作確認済み)"),
 ("Bund 10Y利回り", "Eurex RX1先物→インプライド", "履歴データBUND_FUT(2023-01〜)。スポットBundはBundesbank SDMX BBK01.WT3311"),
 ("JGB 10Y利回り", "財務省 国債金利情報", "履歴データJGB_10Y(ヒストリカル込み)"),
 ("JGB先物", "NEXT FUNDS 10年国債先物ETF(2510.T)", "履歴データJGB_FUT(ETF価格→インプライド利回り換算)"),
 ("EUR/JPYスワップ", "—", "無料の日次APIなし。橙セル手入力 or BBG=BDH()で差し替え"),
]:
    wsAsw.cell(row=r,column=2,value=label).font=Font(size=9,bold=True)
    wsAsw.cell(row=r,column=3,value=src).font=Font(size=9,color=STEEL)
    wsAsw.cell(row=r,column=4,value=note).font=Font(size=9,color="595959")
    for c in range(2,6):wsAsw.cell(row=r,column=c).border=BORDER
    r+=1
print("ASW 構築完了")

# =====================================================================
# ファンディング(履歴データ参照 + レポ想定)
# =====================================================================
ws=wb.create_sheet("ファンディング")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,34,14,44])
title_block(ws,"ファンディング & レポ(最新値 = 履歴データ参照)","NY連銀レートは履歴データ参照。レポGC/SC想定は橙色セルを編集。",4)
ws.append([]);ws.append(["","レート","水準%","備考"])
style_header(ws,3,1,4)
fund_rows=[
 ("SOFR(担保付O/N)",LATEST("SOFR"),"米国の担保付翌日物。資金調達の基準"),
 ("EFFR(無担保O/N)",LATEST("EFFR"),"米国の無担保翌日物"),
 ("TGCR(三方GCレポ)",LATEST("TGCR"),"米国債の一般担保(GC)レポ"),
 ("BGCR(広義GCレポ)",LATEST("BGCR"),"広義の一般担保レポ"),
]
r=4
for name,formula,note in fund_rows:
    ws.cell(row=r,column=2,value=name)
    c=ws.cell(row=r,column=3,value=f"={formula}");c.number_format="0.00"
    ws.cell(row=r,column=4,value=note).font=Font(size=9,color="595959")
    for cc in range(2,5):ws.cell(row=r,column=cc).border=BORDER
    r+=1
r+=1
ws.cell(row=r,column=2,value="レポ資金コスト想定(橙色セルを編集)").font=SUB_FONT
r+=1
ws.cell(row=r,column=3,value="想定%")
style_header(ws,r,2,4,fill=SUB_FILL)
r+=1
repo_assump=[
 ("USD GCレポ(UST買いの資金調達)",GC_USD,"GC","TGCR水準。米国債は一般担保"),
 ("USD SCレポ(特別/CTD)",TGCR-0.15,"SC","特別物はGCより約15bp低い(目分量)"),
 ("JPY GCレポ(JGB買いの資金調達)",GC_JPY,"GC","円GCは約-10bp(目分量、正常化後)"),
 ("JPY SCレポ(JGB特別)",-0.25,"SC","JGB特別は約-25bp(目分量)"),
 ("EUR GCレポ(Bund買いの資金調達)",GC_EUR,"GC","ESTR近傍と想定(目分量)。S7/S8のキャリーに使用"),
]
for name,lvl,typ,note in repo_assump:
    ws.cell(row=r,column=2,value=name)
    c=ws.cell(row=r,column=3,value=lvl);c.number_format="0.00";c.fill=ASSUMP_FILL
    ws.cell(row=r,column=4,value=f"{typ} — {note}").font=Font(size=9,color="595959")
    for cc in range(2,5):ws.cell(row=r,column=cc).border=BORDER
    r+=1
ws.cell(row=r+1,column=2,value="※キャリー=買いは(利回り-レポ)、売りは(レポ-利回り)。S1ベリー買いBFはキャリーがマイナスになりやすい(実務でも同様)。").font=NOTE_FONT

# =====================================================================
# 損益(キャリー内訳)
# =====================================================================
ws=wb.create_sheet("損益")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,30,14,14,40])
title_block(ws,"損益 — キャリー内訳(年率)","マーク損益は分析データ/トレード台帳で日次計算。ここではキャリー(保有収益)の内訳。",5)
ws.append([]);ws.append(["","戦略","キャリー$/年","キャリーbp/年相当","備考"])
style_header(ws,3,1,5)
carry_rows=[
 ("S1 UST 2s10s30s BF",CARRY_S1,"ベリー買いBFはキャリーマイナスが通常(ロールダウン狙い)"),
 ("S2 JGB 2s10s30s BF",CARRY_S2,"ベリー売り=ポジティブキャリー"),
 ("S3 UST-JGB 10年 クロスマーケット",CARRY_S3,"日米キャリー差"),
 ("S4 UST ASWボックス 5s10s",CARRY_S4,"現物ベースの簡易キャリー(ASW水準は『ASW』シート)"),
 ("S5 ドル円オーバーレイ",CARRY_S5,"FXキャリーは金利差(別途)"),
 ("S6 UST 10年 ベーシス",CARRY_S6,"現物キャリー(レポ控除後)"),
 ("S7 Bund-UST 10年 クロスマーケット",CARRY_S7,"Bund買い(負キャリー)+UST売り(負キャリー)。EURレポ想定はファンディングシート"),
 ("S8 JGB-Bund 10年 クロスマーケット",CARRY_S8,"JGB買い(正キャリー)+Bund売り(正キャリー)"),
]
r=4
for name,c,note in carry_rows:
    ws.cell(row=r,column=2,value=name)
    cc=ws.cell(row=r,column=3,value=round(c));cc.number_format="#,##0"
    cc.font=Font(color=GREEN if c>0 else RED,bold=True)
    ws.cell(row=r,column=4,value="")
    ws.cell(row=r,column=5,value=note).font=Font(size=9,color="595959")
    for cc2 in range(2,6):ws.cell(row=r,column=cc2).border=BORDER
    r+=1
ws.cell(row=r,column=2,value="ブック合計").font=BOLD
ct=ws.cell(row=r,column=3,value=f"=SUM(C4:C{r-1})");ct.number_format="#,##0";ct.font=BOLD
for c in range(2,6):ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=LIGHT);ws.cell(row=r,column=c).border=BORDER
ws.cell(row=r+2,column=2,value="※キャリーは現在利回り×レポ想定から算出した年率概算。マーク損益(価格変動)は分析データが日次で計算し仮NAVに反映。").font=NOTE_FONT

# =====================================================================
# ポートフォリオ(建玉)
# =====================================================================
ws=wb.create_sheet("ポートフォリオ")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,26,10,14,26,10,13,13,14,40])
title_block(ws,"モデルポートフォリオ — 建玉","各戦略のレッグとDV01。全戦略DV01ニュートラル。",10)
ws.append([])
hdr=["","戦略","銘柄","説明","方向","ノショナル(百万$)","利回り%","DV01/$1mm","レッグDV01($/bp)","ヘッジ役割"]
ws.append(hdr)
style_header(ws,3,1,10)
def legrow(strat,iid,desc,direction,notional,y,d01mm,hedge):
    return [strat,iid,desc,"買い" if direction>0 else "売り",notional,y,d01mm,direction*notional*d01mm,hedge]
n10=S1_DV01/d10;n2=S1_DV01/2/d2;n30=S1_DV01/2/d30
jn10=S2_DV01/j10;jn2=S2_DV01/2/j2;jn30=S2_DV01/2/j30
xn_jgb=S3_DV01/j10;xn_ust=S3_DV01/d10
an5=S4_DV01/d5;an10=S4_DV01/d10
bn_cash=S6_DV01/d10
n7_long=S7_DV01/b10;n7_short=S7_DV01/d10
n8_long=S8_DV01/g10;n8_short=S8_DV01/b10
positions=[
 legrow("S1 UST 2s10s30s BF","UST10Y","米10年(ベリー)",+1,round(n10,1),UST_Y["10Y"],round(d10,0),"リスク:ベリー買い"),
 legrow("S1 UST 2s10s30s BF","UST2Y","米2年(フロントウィング)",-1,round(n2,1),UST_Y["2Y"],round(d2,0),"ヘッジ:短期金利"),
 legrow("S1 UST 2s10s30s BF","UST30Y","米30年(バックウィング)",-1,round(n30,1),UST_Y["30Y"],round(d30,0),"ヘッジ:超長期金利"),
 legrow("S2 JGB 2s10s30s BF","JGB10Y","日本10年(ベリー)",-1,round(jn10,1),JGB_Y["10Y"],round(j10,0),"リスク:ベリー売り"),
 legrow("S2 JGB 2s10s30s BF","JGB2Y","日本2年(フロントウィング)",+1,round(jn2,1),JGB_Y["2Y"],round(j2,0),"ヘッジ"),
 legrow("S2 JGB 2s10s30s BF","JGB30Y","日本30年(バックウィング)",+1,round(jn30,1),JGB_Y["30Y"],round(j30,0),"ヘッジ"),
 legrow("S3 UST-JGB 10年 クロスマーケット","JGB10Y","日本10年",+1,round(xn_jgb,1),JGB_Y["10Y"],round(j10,0),"リスク:JGB買い"),
 legrow("S3 UST-JGB 10年 クロスマーケット","UST10Y","米10年",-1,round(xn_ust,1),UST_Y["10Y"],round(d10,0),"ヘッジ:UST売り"),
 legrow("S4 UST ASWボックス 5s10s","UST5Y","米5年現物+固定受け",+1,round(an5,1),UST_Y["5Y"],round(d5,0),"リスク:ASW受け"),
 legrow("S4 UST ASWボックス 5s10s","UST10Y","米10年現物+固定払い",-1,round(an10,1),UST_Y["10Y"],round(d10,0),"ヘッジ:ASW払い"),
 legrow("S5 ドル円オーバーレイ","USDJPY","ドル円売り(円買い)",-1,S5_NOT,USDJPY,0.0,"分散:低相関FX"),
 legrow("S6 UST 10年 ベーシス","UST10Y","米10年現物(CTD相当)",+1,round(bn_cash,1),UST_Y["10Y"],round(d10,0),"リスク:現物買い"),
 legrow("S6 UST 10年 ベーシス","ZN1","ZN 10年債先物",-1,float(bn_fut),0.0,ZN_DV01,"ヘッジ:先物売り"),
 legrow("S7 Bund-UST 10年 クロスマーケット","BUND_FUT","Euro-Bund 10Y先物(インプライド)",+1,round(n7_long,1),BUND_Y,round(b10,0),"リスク:Bund買い"),
 legrow("S7 Bund-UST 10年 クロスマーケット","UST10Y","米10年",-1,round(n7_short,1),UST_Y["10Y"],round(d10,0),"ヘッジ:UST売り"),
 legrow("S8 JGB-Bund 10年 クロスマーケット","JGB_FUT","NEXT FUNDS 10年国債先物(2510.T)",+1,round(n8_long,1),JGBF_Y,round(g10,0),"リスク:JGB先物買い"),
 legrow("S8 JGB-Bund 10年 クロスマーケット","BUND_FUT","Euro-Bund 10Y先物(インプライド)",-1,round(n8_short,1),BUND_Y,round(b10,0),"ヘッジ:Bund売り"),
]
r=4
for p in positions:
    ws.cell(row=r,column=2,value=p[0])
    ws.cell(row=r,column=3,value=p[1]).font=BOLD
    ws.cell(row=r,column=4,value=p[2])
    dcell=ws.cell(row=r,column=5,value=p[3])
    dcell.font=Font(bold=True,color=GREEN if p[3]=="買い" else RED)
    ws.cell(row=r,column=6,value=p[4]).number_format="#,##0.0"
    ws.cell(row=r,column=7,value=p[5]).number_format="0.000"
    ws.cell(row=r,column=8,value=p[6]).number_format="#,##0"
    cl=ws.cell(row=r,column=9,value=round(p[7],0));cl.number_format="#,##0"
    cl.font=Font(color=GREEN if p[7]>0 else RED)
    ws.cell(row=r,column=10,value=p[8]).font=Font(size=9)
    for c in range(2,11):
        ws.cell(row=r,column=c).border=BORDER
        if r%2==0:ws.cell(row=r,column=c).fill=ALT_FILL
    r+=1
tot_not=sum(p[4] for p in positions)
ws.cell(row=r,column=2,value="グロスノショナル合計").font=BOLD
ws.cell(row=r,column=6,value=round(tot_not,1)).number_format="#,##0.0"
for c in range(2,11):ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=LIGHT);ws.cell(row=r,column=c).border=BORDER
ws.cell(row=r+2,column=2,value="※各戦略はDV01ニュートラル(リスクレッグのDV01をヘッジレッグで相殺)。損益は金利水準でなく相対価値から。").font=NOTE_FONT
ws.freeze_panes="A4"

# =====================================================================
# 戦略解説
# =====================================================================
ws=wb.create_sheet("戦略解説")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,22,38,38,38,14])
title_block(ws,"戦略解説 — 仕組み・エントリータイミング","何を買い何を売るか / なぜ儲かるか / ヘッジ / いつ入るか",6)
ws.append([])
ws.append(["","戦略","① 何をやるか","② なぜ儲かるか","③ ヘッジ+エントリータイミング","キャリー$/年"])
style_header(ws,3,1,6)
strat_cards=[
 ("S1 UST 2s10s30s BF",
  "米10年(ベリー)買い、米2年+米30年(ウィング)売り。",
  "カーブのお腹(10年)が割安と見て、バタフライ的にフラット化すると利益。金利の上下には依存しない。",
  "ウィング売りのDV01でベリー買いを相殺(平行シフト除去)。エントリーはバタフライスプレッドのZスコアが-2σ(割安)でロングBF、+2σでショートBF。",
  round(CARRY_S1)),
 ("S2 JGB 2s10s30s BF",
  "日本10年(ベリー)売り、日本2年+30年(ウィング)買い。",
  "日銀正常化でJGBカーブがベアスティープ化、10年がアンダーパフォームすると見る。",
  "ウィング買いでベリー売りのDV01を相殺。エントリーはJGBバタフライスプレッドの±2σ。",
  round(CARRY_S2)),
 ("S3 UST-JGB 10年 クロスマーケット",
  "日本10年買い、米10年売り(DV01一致)。",
  "米-日10年ベーシス(約190bp)が広すぎ、縮小(JGB相対アウトパフォーム)を狙う。",
  "UST売りが米水準、JGB買いが日水準を除去し純粋ベーシスに。エントリーはベーシスが+2σ(拡大しすぎ)で縮小狙い。",
  round(CARRY_S3)),
 ("S4 UST ASWボックス 5s10s",
  "米5年現物+固定受け(ASW) vs 米10年現物+固定払い(ASW)。",
  "ネガティブキャリーの側をショート、ポジティブキャリーの側をロングにして、5s10sのスワップスプレッド形状を狙う。",
  "受け/払いでDV01一致、金利水準を除去。スワップデータ未取得のため5s10sスロープで代用(BBGでASW取得後に差し替え)。",
  round(CARRY_S4)),
 ("S5 ドル円オーバーレイ",
  "ドル円売り(円買い)。",
  "G4金利と低相関のFXで、ブック全体の残存リスクと日本固有テールを分散。",
  "DV01ヘッジではなく低相関分散。金利戦略と独立した収益源。",
  round(CARRY_S5)),
 ("S6 UST 10年 ベーシス",
  "米10年現物買い、ZN先物売り = ベーシス買い。Capulaの代名詞。",
  "現先ベーシスが受渡に向けて収斂するのを狙い、レポ調達でキャリーを稼ぐ。",
  "先物売りが現物のDV01を相殺し水準除去。【注意】レポレバレッジのため2020年3月のようなベーシス急拡大が最大リスク。",
  round(CARRY_S6)),
 ("S7 Bund-UST 10年 クロスマーケット",
  "Bund先物インプライド利回り(履歴データBUND_FUT)買い、米10年売り(DV01一致)。",
  "Bund-USTの10年ベーシスが極端に乖離したら収斂を狙う。Bund買いが欧元金利水準、UST売りが米金利水準を除去し純粋なクロスマーケットベーシスに。",
  "エントリーはベーシスが+2σ(拡大しすぎ)で縮小狙い。キャリーはBund買い(負)+UST売り(負)で両ネガティブ。",
  round(CARRY_S7)),
 ("S8 JGB-Bund 10年 クロスマーケット",
  "JGB先物インプライド(2510 ETF換算)買い、Bund先物インプライド売り(DV01一致)。",
  "JGB-Bundの10年ベーシスが極端に乖離したら収斂を狙う。JGB側が日本金利、Bund側がユーロ金利。",
  "エントリーはベーシスが+2σ(拡大しすぎ)で縮小狙い。キャリーはJGB買い(正)+Bund売り(正)で両ポジティブ。",
  round(CARRY_S8)),
]
r=4
for name,pos,thesis,hedge,carry in strat_cards:
    ws.cell(row=r,column=2,value=name).font=BOLD
    ws.cell(row=r,column=3,value=pos).alignment=TOPLEFT
    ws.cell(row=r,column=4,value=thesis).alignment=TOPLEFT
    ws.cell(row=r,column=5,value=hedge).alignment=TOPLEFT
    cc=ws.cell(row=r,column=6,value=carry);cc.number_format="#,##0"
    cc.font=Font(color=GREEN if carry>0 else RED)
    for c in range(2,7):ws.cell(row=r,column=c).border=BORDER
    ws.row_dimensions[r].height=90
    r+=1
ws.cell(row=r+1,column=2,value="※エントリータイミングは『エントリーシグナル』シートのZスコア±2σを参照。実運用ではキャリー・ファンディング・イベントリスクを重ねて最終判断。").font=NOTE_FONT

# =====================================================================
# 銘柄マスター
# =====================================================================
ws=wb.create_sheet("銘柄マスター")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,14,30,10,8,12,12,14,26])
title_block(ws,"銘柄マスター","ブックで使うベンチマーク銘柄 + BBGティッカー(BBG使う場合)",9)
ws.append([]);ws.append(["","銘柄ID","説明","市場","年限","利回り%","修正デュレーション","DV01/$1mm","BBGティッカー(要確認)"])
style_header(ws,3,1,9)
instruments=[
 ("UST2Y","米国債 2年(最新回)","UST","2Y",UST_Y["2Y"],"USGG2YR Index"),
 ("UST5Y","米国債 5年(最新回)","UST","5Y",UST_Y["5Y"],"USGG5YR Index"),
 ("UST10Y","米国債 10年(最新回)","UST","10Y",UST_Y["10Y"],"USGG10YR Index"),
 ("UST30Y","米国債 30年(最新回)","UST","30Y",UST_Y["30Y"],"USGG30YR Index"),
 ("JGB2Y","日本国債 2年(最新回)","JGB","2Y",JGB_Y["2Y"],"GJGB2 Index"),
 ("JGB10Y","日本国債 10年(最新回)","JGB","10Y",JGB_Y["10Y"],"GJGB10 Index"),
 ("JGB20Y","日本国債 20年(最新回)","JGB","20Y",JGB_Y["20Y"],"GJGB20 Index"),
 ("JGB30Y","日本国債 30年(最新回)","JGB","30Y",JGB_Y["30Y"],"GJGB30 Index"),
 ("ZN1","米10年債先物(ZN)","先物","10Y",None,"ZN1 Comdty"),
 ("JB1","日本国債10年先物(JB)","先物","10Y",None,"JB1 Comdty"),
 ("RX1","Euro-Bund 10年先物","先物","10Y",BUND_Y,"RX1 Comdty"),
 ("JGBF","NEXT FUNDS 10年国債先物ETF(2510.T)","先物","10Y",JGBF_Y,"2510 JP Equity"),
 ("USDJPY","ドル円スポット","FX","-",None,"USDJPY Curncy"),
]
r=4
for iid,desc,mkt,ten,y,bbg in instruments:
    if y is not None:
        n=float(ten.rstrip("Y"));md=par_mod_dur(n,y);d01=dv01_per_1mm(md)
    else:
        md=None;d01=None
    ws.cell(row=r,column=2,value=iid).font=BOLD
    ws.cell(row=r,column=3,value=desc)
    ws.cell(row=r,column=4,value=mkt)
    ws.cell(row=r,column=5,value=ten)
    ws.cell(row=r,column=6,value=y).number_format="0.000"
    ws.cell(row=r,column=7,value=round(md,2) if md else None).number_format="0.00"
    ws.cell(row=r,column=8,value=round(d01,0) if d01 else None).number_format="#,##0"
    ws.cell(row=r,column=9,value=bbg).font=Font(color=STEEL)
    for c in range(2,10):
        ws.cell(row=r,column=c).border=BORDER
        if r%2==0:ws.cell(row=r,column=c).fill=ALT_FILL
    r+=1
ws.cell(row=r+1,column=2,value="※BBGを使う場合、ティッカー列を参考に=BDP()を貼ってください(このブック自体にはBDP式は入っていません)。").font=NOTE_FONT

# =====================================================================
# はじめに
# =====================================================================
ws=wb.create_sheet("はじめに",0)
ws.sheet_view.showGridLines=False
set_widths(ws,[3,118])
title_block(ws,"Capula型 債券レラティブバリュー(RV)モデル — ファンド管理版","v3.0 / マクロ(VBA)でデータ更新 / 基準日 "+ASOF,2)
readme=[
("これは何か",
 "ヘッジファンドCapula Investment Managementのスタイルを模した、債券レラティブバリュー(相対価値)戦略の"
 "モデルポートフォリオ+ファンド管理フレームです。8戦略をDV01ヘッジ済み・キャリー/レポコスト込みで組み、"
 "仮NAV・シャーレシオ・最大DD・VaR・エントリーシグナル・トレード台帳まで備えています。"),
("【重要】データ更新はマクロ(VBA)で行います",
 "ダッシュボードの緑の【データ更新】ボタンを押すと、VBAが米財務省/日本財務省/NY連銀/Yahoo/Pensford(ASWスワップ)から"
 "ヒストリカルデータ込みで再取得し、『履歴データ』シートを書き換えます。全シートは履歴データを数式参照しているので"
 "自動で再計算されます。BDP関数は入っていません(BBGを使う場合はあなたが貼ってください)。"),
("【重要】マクロの有効化が必要",
 ".xlsm形式です。開いたときセキュリティ警告が出たら【コンテンツの有効化】を押してください。"
 "VBAが動かないとデータ更新ボタンが使えません。"),
("シートの見方",
 "①ダッシュボード=仮NAV・リスク指標・更新ボタン → ②戦略解説=各戦略の仕組み → ③エントリーシグナル=いつ入るか(Zスコア±2σ+チャート) → "
 "④トレード台帳=戦略別収益表+オープン/クローズドトレード → ⑤ポートフォリオ=建玉 → ⑥分析データ=日次P&L→仮NAVの数式エンジン → "
 "⑦履歴データ=唯一のデータ源(VBAが更新) → ⑧イールドカーブ/ASW/ファンディング/損益/銘柄マスター。"),
("仮NAVの仕組み",
 "2023-01からのバックテストです。各戦略の日次損益をスプレッド変化×DV01で計算し、キャリーを足して累積。"
 "基数は$100mm。実運用のNAVではなく、モデルの感応度・リスクを測るためのものです。"),
("ヘッジの考え方",
 "各戦略はDV01ニュートラル。リスクレッグの金利方向リスクをヘッジレッグで相殺し、損益は相対価値(カーブ/ベーシス/スプレッド)から。"
 "残存のブック方向リスクは低相関のドル円オーバーレイで分散。"),
("ASW(資産スワップ)について",
 "『ASW』シートに米国債のASWカーブ(2Y-10Y)を表示。スワップレートはPensfordの公開API(https://pensford.com/api/live-rates)から"
 "VBA更新時に自動取得します(Treasury利回り−SOFRスワップ)。Bund/JGBのスワップは無料APIが無いため、"
 "現状は先物インプライド利回りと手入力(橙色セル)のEUR/JPYスワップから算出。BBGがあればそのまま貼り替えできます。"),
("S4 ASWボックスのキャリーロジック",
 "ASWボックスはネガティブキャリーの側をショート、ポジティブキャリーの側をロングにします。"
 "ただしスワップの日次履歴が無料では取得できないため、バックテストのP&Lエンジンは5s10sスロープで代用し、"
 "現在のASW水準だけ『ASW』シートでモニターします。"),
("次のフェーズ",
 "リスク管理の高度化(ストレステスト具体化、シナリオショック、グリークス、リミット管理)、"
 "EUR/JPYスワップレートの取得(有料データかBBG)、ASW履歴の蓄積(毎日VBA更新でスナップショットを積む)など。"),
]
r=4
for head,body in readme:
    ws.cell(row=r,column=2,value=head).font=SUB_FONT
    r+=1
    c=ws.cell(row=r,column=2,value=body);c.alignment=TOPLEFT
    ws.row_dimensions[r].height=66
    r+=2

# =====================================================================
# リスク(Phase 2: VaR比較 / シナリオストレス / グリークス / リミット / デレバレッジ)
# =====================================================================
from openpyxl.formatting.rule import CellIsRule
ws=wb.create_sheet("リスク")
ws.sheet_view.showGridLines=False
set_widths(ws,[3,28,36,12,12,12,12,12,12,12,12,12])
title_block(ws,"リスク管理 — VaR・ストレステスト・グリークス・リミット","Capula型: 日次VaR+ストレステスト+ポジションリミット+流動性監視。全数式は分析データ参照なのでVBA更新で自動再計算。",12)

WARN_FILL=PatternFill("solid",fgColor="FFEB9C")

# ---- ■ VaR比較 ----
ws.cell(row=4,column=2,value="■ VaR比較(仮NAVベース)").font=SUB_FONT
ws.cell(row=5,column=2,value="指標");ws.cell(row=5,column=3,value="値");ws.cell(row=5,column=4,value="説明")
style_header(ws,5,2,4)
var_rows=[
 ("ヒストリカルVaR99(1日,$mm)",f"=-PERCENTILE(分析データ!$AD$2:$AD${CAP_LAST},0.01)","0.00","日次損益の実分布の1%タイル。分布形状を仮定しない"),
 ("パラメトリックVaR99(1日,$mm)",f"=INDEX(分析データ!$AE:$AE,{M_LATEST})*(-(AVERAGE(分析データ!$AF$2:$AF${CAP_LAST})-2.3263*STDEV(分析データ!$AF$2:$AF${CAP_LAST})))","0.00","正規分布仮定(z99=2.3263)。ヒストリカルと乖離が大きい=裾が厚い"),
 ("パラメトリックVaR99(10日,$mm)",f"=C7*SQRT(10)","0.00","√10スケーリング。機関投資家レポート標準"),
 ("最悪60日累積損失($mm)",f"=-MIN($O$4:$O${3+CAP_LAST-60})","0.00","ローリング60営業日P&Lの最悪値(ストレスVaR相当)。右端ヘルパー列参照"),
]
r=6
for name,formula,fmt,desc in var_rows:
    ws.cell(row=r,column=2,value=name)
    c=ws.cell(row=r,column=3,value=formula);c.number_format=fmt;c.font=BOLD
    ws.cell(row=r,column=4,value=desc).font=Font(size=9,color="595959")
    for cc in range(2,5):ws.cell(row=r,column=cc).border=BORDER
    r+=1

# ---- ■ シナリオショックシミュレータ(編集可能) ----
ws.cell(row=11,column=2,value="■ シナリオショックシミュレータ(橙色セルを編集)").font=SUB_FONT
shocks=[
 ("ΔUST 2年(bp)",0,"0.0"),
 ("ΔUST 5年(bp)",0,"0.0"),
 ("ΔUST 10年(bp)",50,"0.0"),
 ("ΔUST 30年(bp)",50,"0.0"),
 ("ΔJGB 2年(bp)",0,"0.0"),
 ("ΔJGB 10年(bp)",0,"0.0"),
 ("ΔJGB 30年(bp)",0,"0.0"),
 ("ΔBund 10年(bp)",0,"0.0"),
 ("ΔUSDJPY(%)",0.0,"0.0%"),
 ("ΔUST現先ベーシス(bp)",0,"0.0"),
]
ws.cell(row=12,column=2,value="ショック入力");ws.cell(row=12,column=3,value="値")
style_header(ws,12,2,3,fill=SUB_FILL)
sr=13
for name,val,fmt in shocks:
    ws.cell(row=sr,column=2,value=name)
    c=ws.cell(row=sr,column=3,value=val);c.fill=ASSUMP_FILL;c.number_format=fmt
    for cc in (2,3):ws.cell(row=sr,column=cc).border=BORDER
    sr+=1
# C13..C22 がショックセル(C20=ΔBund 10年、C21=ΔUSDJPY、C22=Δ現先ベーシス)
ws.cell(row=12,column=5,value="戦略");ws.cell(row=12,column=6,value="損益インパクト$")
style_header(ws,12,5,6,fill=SUB_FILL)
sim_impacts=[
 ("S1 UST 2s10s30s BF",f"=-{S1_DV01}*(C15-(C13+C16)/2)"),
 ("S2 JGB 2s10s30s BF",f"={S2_DV01}*(C18-(C17+C19)/2)"),
 ("S3 UST-JGB クロスマーケット",f"=-{S3_DV01}*(C15-C18)"),
 ("S4 UST 5s10s",f"=-{S4_DV01}*(C15-C14)"),
 ("S5 ドル円オーバーレイ",f"=-{S5_NOT}*1000000*C21"),
 ("S6 UST 10年 ベーシス",f"=-{S6_DV01}*C22"),
 ("S7 Bund-UST クロスマーケット",f"=-{S7_DV01}*(C20-C15)"),
 ("S8 JGB-Bund クロスマーケット",f"=-{S8_DV01}*(C18-C20)"),
]
ir=13
for name,formula in sim_impacts:
    ws.cell(row=ir,column=5,value=name)
    c=ws.cell(row=ir,column=6,value=formula);c.number_format="#,##0"
    for cc in (5,6):ws.cell(row=ir,column=cc).border=BORDER
    ir+=1
ws.cell(row=ir,column=5,value="ブック合計").font=BOLD
tc=ws.cell(row=ir,column=6,value=f"=SUM(F13:F{ir-1})");tc.number_format="#,##0";tc.font=BOLD
for cc in (5,6):ws.cell(row=ir,column=cc).border=BORDER;ws.cell(row=ir,column=cc).fill=PatternFill("solid",fgColor=LIGHT)
ir+=1
ws.cell(row=ir,column=5,value="NAVインパクト%").font=BOLD
nc=ws.cell(row=ir,column=6,value=f"=F{ir-1}/1000000/INDEX(分析データ!$AE:$AE,{M_LATEST})");nc.number_format="0.00%";nc.font=BOLD
for cc in (5,6):ws.cell(row=ir,column=cc).border=BORDER

# ---- ■ 歴史的シナリオ(固定ショック) ----
pr=25
ws.cell(row=pr-1,column=2,value="■ 歴史的シナリオ・プレセット(過去の危機を現在ポジションに適用)").font=SUB_FONT
hdr=["","シナリオ","ショック内容","S1","S2","S3","S4","S5","S6","S7","S8","ブック合計$","NAV%"]
for i,h in enumerate(hdr):ws.cell(row=pr,column=1+i,value=h)
style_header(ws,pr,1,13)
presets=[
 ("2020年3月 ダッシュフォーキャッシュ","ベーシス急拡大+全面金利上昇+円高",30,40,50,40,0,0,30,0,-0.05,60),
 ("2022年9月 英LDI危機","超長期中心の投げ売り(米債プロキシ)",50,55,60,80,0,0,50,0,0.0,20),
 ("2023年3月 SVB","急速なブルスティープナー",-40,-25,-10,0,0,0,-30,0,0.0,0),
 ("2022年12月 日銀YCC変更","JGB10年+25bp・円高",0,0,0,0,5,25,0,15,-0.03,0),
 ("ベアスティープナー","UST 10年+50/30年+60",0,25,50,60,0,0,40,0,0.0,0),
 ("ベアフラットナー","UST 2年+50/10年+30",50,40,30,20,0,0,30,0,0.0,0),
 ("円高ショック(介入)","USDJPY -8%",0,0,0,0,0,0,0,0,-0.08,0),
 ("現先ベーシス急拡大","S6の殺し手: ベーシス+50bp",0,0,0,0,0,0,0,0,0.0,50),
]
r=pr+1
for name,desc,du2,du5,du10,du30,dj2,dj10,dj30,dbund,dfx,dbasis in presets:
    ws.cell(row=r,column=2,value=name)
    ws.cell(row=r,column=3,value=desc).font=Font(size=9)
    ws.cell(row=r,column=4,value=f"=-{S1_DV01}*({du10}-({du2}+{du30})/2)").number_format="#,##0"
    ws.cell(row=r,column=5,value=f"={S2_DV01}*({dj10}-({dj2}+{dj30})/2)").number_format="#,##0"
    ws.cell(row=r,column=6,value=f"=-{S3_DV01}*({du10}-{dj10})").number_format="#,##0"
    ws.cell(row=r,column=7,value=f"=-{S4_DV01}*({du10}-{du5})").number_format="#,##0"
    ws.cell(row=r,column=8,value=f"=-{S5_NOT}*1000000*{dfx}").number_format="#,##0"
    ws.cell(row=r,column=9,value=f"=-{S6_DV01}*{dbasis}").number_format="#,##0"
    ws.cell(row=r,column=10,value=f"=-{S7_DV01}*({dbund}-{du10})").number_format="#,##0"
    ws.cell(row=r,column=11,value=f"=-{S8_DV01}*({dj10}-{dbund})").number_format="#,##0"
    ws.cell(row=r,column=12,value=f"=SUM(D{r}:K{r})").number_format="#,##0"
    ws.cell(row=r,column=13,value=f"=L{r}/1000000/INDEX(分析データ!$AE:$AE,{M_LATEST})").number_format="0.00%"
    for c in range(2,14):ws.cell(row=r,column=c).border=BORDER
    ws.cell(row=r,column=12).font=BOLD
    r+=1

# ---- ■ グリークス ----
gr=r+1
ws.cell(row=gr,column=2,value="■ グリークス(ネットエクスポージャー)").font=SUB_FONT
gr+=1
ws.cell(row=gr,column=2,value="項目");ws.cell(row=gr,column=3,value="現在値");ws.cell(row=gr,column=4,value="備考")
style_header(ws,gr,2,4)
greeks=[
 ("ネットDV01 UST($/bp)",-(S3_DV01+S7_DV01),"#,##0","S3のUST売りとS7のUST売りが残存。S1/S4/S6はUST内でニュートラル"),
 ("ネットDV01 JGB($/bp)",(S3_DV01+S8_DV01),"#,##0","S3のJGB買いとS8のJGB買いが残存。S2はJGB内でニュートラル"),
 ("ネットDV01 Bund($/bp)",(S7_DV01-S8_DV01),"#,##0","S7のBund買いとS8のBund売りが同額で相殺(ゼロ)"),
 ("FXデルタ(百万$)",-S5_NOT,"#,##0","ドル売り・円買い(S5)。低相関分散が目的"),
 ("日次キャリー($/日)",CARRY_TOTAL/252,"#,##0","全戦略キャリー年率÷252"),
 ("グロスノショナル(百万$)",tot_not,"#,##0.0","全レッグ合計"),
 ("コンベクシティ","各戦略DV01ニュートラルのため二次リスクは小さい","@","大きなショックは上記シナリオ表で評価"),
]
r=gr+1
for name,val,fmt,note in greeks:
    ws.cell(row=r,column=2,value=name)
    c=ws.cell(row=r,column=3,value=val);c.number_format=fmt;c.font=BOLD
    ws.cell(row=r,column=4,value=note).font=Font(size=9,color="595959")
    for cc in range(2,5):ws.cell(row=r,column=cc).border=BORDER
    r+=1

# ---- ■ リミット管理 ----
lr=r+1
ws.cell(row=lr,column=2,value="■ リミット管理(橙色=リミット入力)").font=SUB_FONT
lr+=1
ws.cell(row=lr,column=2,value="項目");ws.cell(row=lr,column=3,value="現在値");ws.cell(row=lr,column=4,value="リミット");ws.cell(row=lr,column=5,value="利用率");ws.cell(row=lr,column=6,value="状態")
style_header(ws,lr,2,6)
lim_rows=[
 ("ブックVaR99(1日,$mm)",f"=-PERCENTILE(分析データ!$AD$2:$AD${CAP_LAST},0.01)","0.00",3.0,"=C{r}/D{r}"),
 ("最大戦略DV01($/bp)",max(S1_DV01,S2_DV01,S3_DV01,S4_DV01,S6_DV01,S7_DV01,S8_DV01),"#,##0",100000,"=C{r}/D{r}"),
 ("ネットDV01 UST($/bp)",-(S3_DV01+S7_DV01),"#,##0",150000,"=ABS(C{r})/D{r}"),
 ("ネットDV01 JGB($/bp)",(S3_DV01+S8_DV01),"#,##0",150000,"=ABS(C{r})/D{r}"),
 ("ネットDV01 Bund($/bp)",(S7_DV01-S8_DV01),"#,##0",75000,"=ABS(C{r})/D{r}"),
 ("FXデルタ(百万$)",-S5_NOT,"#,##0",200,"=ABS(C{r})/D{r}"),
 ("グロスレバレッジ(倍)",f"={tot_not:.0f}/INDEX(分析データ!$AE:$AE,{M_LATEST})","0.00",8.0,"=C{r}/D{r}"),
]
r=lr+1
status_cells=[]
for name,val,fmt,limit,utilf in lim_rows:
    ws.cell(row=r,column=2,value=name)
    c=ws.cell(row=r,column=3,value=val);c.number_format=fmt
    d=ws.cell(row=r,column=4,value=limit);d.fill=ASSUMP_FILL;d.number_format=fmt
    ws.cell(row=r,column=5,value=utilf.format(r=r)).number_format="0%"
    ws.cell(row=r,column=6,value=f'=IF(E{r}>=1,"限度超過",IF(E{r}>=0.8,"警告","OK"))')
    for cc in range(2,7):ws.cell(row=r,column=cc).border=BORDER
    status_cells.append(f"F{r}")
    r+=1
# DDトリガー(利用率なし)
ws.cell(row=r,column=2,value="現在DD%")
ws.cell(row=r,column=3,value=f"=MIN(分析データ!$AH$2:$AH${CAP_LAST})").number_format="0.0%"
d=ws.cell(row=r,column=4,value=-0.20);d.fill=ASSUMP_FILL;d.number_format="0%"
ws.cell(row=r,column=5,value="")
ws.cell(row=r,column=6,value=f'=IF(C{r}<=D{r},"STOP: デレバレッジ",IF(C{r}<=D{r}/2,"警告","OK"))')
for cc in range(2,7):ws.cell(row=r,column=cc).border=BORDER
status_cells.append(f"F{r}")
r+=1
# 条件付き書式: 状態列
ws.conditional_formatting.add(f"F{lr+1}:F{r-1}",
    CellIsRule(operator="equal",formula=['"限度超過"'],fill=BAD_FILL))
ws.conditional_formatting.add(f"F{lr+1}:F{r-1}",
    CellIsRule(operator="equal",formula=['"警告"'],fill=WARN_FILL))
ws.conditional_formatting.add(f"F{lr+1}:F{r-1}",
    CellIsRule(operator="equal",formula=['"STOP: デレバレッジ"'],fill=BAD_FILL))

# ---- ■ ファンディングストレス監視 ----
fr=r+1
ws.cell(row=fr,column=2,value="■ ファンディングストレス監視(2020年3月の教訓)").font=SUB_FONT
fr+=1
fund_rows=[
 ("SOFR最新(%)",f"=INDEX(履歴データ!$K:$K,{M_LATEST})","0.00"),
 ("SOFR 20日平均(%)",f"=AVERAGE(OFFSET(履歴データ!$K$1,{M_LATEST}-20,0,20,1))","0.00"),
 ("SOFRスパイク(bp)",f"=(C{fr}-C{fr+1})*100","0.0"),
 ("TGCR最新(GCレポ,%)",f"=INDEX(履歴データ!$M:$M,{M_LATEST})","0.00"),
]
r=fr
for name,formula,fmt in fund_rows:
    ws.cell(row=r,column=2,value=name)
    c=ws.cell(row=r,column=3,value=formula);c.number_format=fmt
    for cc in (2,3):ws.cell(row=r,column=cc).border=BORDER
    r+=1
ws.cell(row=r,column=2,value="ファンディング状態")
ws.cell(row=r,column=3,value=f'=IF(C{fr+2}>25,"警告: ファンディングストレス","正常")')
ws.conditional_formatting.add(f"C{r}",CellIsRule(operator="equal",formula=['"警告: ファンディングストレス"'],fill=WARN_FILL))
for cc in (2,3):ws.cell(row=r,column=cc).border=BORDER
r+=2

# ---- ■ デレバレッジルール ----
ws.cell(row=r,column=2,value="■ デレバレッジルール(2020年3月 ダッシュフォーキャッシュの教訓)").font=SUB_FONT
r+=1
delev=[
 "DD -10% → 新規エントリー停止、全ポジションのリスクレビュー。",
 "DD -15% → 全戦略を30%サイズ削減。",
 "DD -25%(STOP) → 50%削減 + S6(ベーシス)を最優先で解消(レポレバレッジが最も高い)。",
 "SOFRスパイク >25bp → ファンディングストレスと判定し、S6のレバレッジを先に落とす。",
 "実績: 2020年3月のベーシス爆発でCapula -$572mm、ExodusPoint -$360mm、Millennium -$100mm(WSJ/Bloomberg)。",
 "レバレッジの高いベーシストレードは、ベーシス拡大+マージンコールの連鎖で殺される。流動性監視が最優先。",
]
for t in delev:
    ws.cell(row=r,column=2,value="・"+t).alignment=LEFT
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    r+=1

# ---- ヘルパー列O: ローリング60日ブックP&L($mm) ----
ws.cell(row=3,column=15,value="60日ローリングP&L$mm(ヘルパー)").font=NOTE_FONT
for k in range(0,CAP_LAST-60):
    row=4+k
    ws.cell(row=row,column=15,value=f'=IF(COUNT(OFFSET(分析データ!$AC$2,{k},0,60,1))<60,"",SUM(OFFSET(分析データ!$AC$2,{k},0,60,1))/1000000)')
    ws.cell(row=row,column=15).number_format="0.00"
ws.column_dimensions["O"].width=12
print("リスク 構築完了")

# =====================================================================
# シート順 + 保存
# =====================================================================
order=["はじめに","ダッシュボード","戦略解説","エントリーシグナル","シグナル検証","トレード台帳","ポートフォリオ",
       "分析データ","履歴データ","イールドカーブ","ASW","ASWデータ","ファンディング","損益","銘柄マスター","リスク"]
wb._sheets.sort(key=lambda s:order.index(s.title))
wb.active=0

# ---- bisect用: KEEP_SHEETS環境変数で先頭Nシートのみ残す(デバッグ) ----
import os as _osenv
_keep=_osenv.environ.get("KEEP_SHEETS")
if _keep:
    keep=order[:int(_keep)]
    for _s in list(wb.worksheets):
        if _s.title not in keep:
            wb.remove(_s)

# ---- 空文字列セルを除去(空のinlineStrはOOXML不正でExcel修復の原因) ----
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _cell in _row:
            if _cell.value == "":
                _cell.value = None

wb.save(OUT)

# ---- 後処理: openpyxlが書く絶対パスTargetを相対パスに修正(Excel修復プロンプト対策) ----
import zipfile, posixpath, shutil, os as _os
def fix_rels(path):
    tmp=path+".tmp"
    with zipfile.ZipFile(path,"r") as zin, zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data=zin.read(item.filename)
            if item.filename.endswith(".rels"):
                base=posixpath.dirname(posixpath.dirname(item.filename))  # dir containing the part
                txt=data.decode("utf-8")
                import re as _re
                def _rel(m):
                    tgt=m.group(1)
                    if tgt.startswith("/"):
                        rel=posixpath.relpath(tgt.lstrip("/"), base if base else ".")
                        return f'Target="{rel}"'
                    return m.group(0)
                txt=_re.sub(r'Target="([^"]+)"',_rel,txt)
                data=txt.encode("utf-8")
            zout.writestr(item,data)
    _os.replace(tmp,path)
fix_rels(OUT)
print("rels修正済み")

print("保存:",OUT)
print("シート:",[s.title for s in wb.worksheets])
